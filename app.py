"""
Fee Defaulter Report Generator — Streamlit Edition
"""
from __future__ import annotations

import bcrypt
import base64
import csv
import hashlib
import hmac
import io
import json
import os
import re
import secrets
import time
import zipfile
from datetime import datetime
from pathlib import Path
from urllib.parse import quote

import streamlit as st

st.set_page_config(
    page_title="Fee Defaulter Report Generator",
    page_icon="📋",
    layout="wide",
    initial_sidebar_state="expanded",
)

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    st.error("Missing required package: openpyxl. Run: pip install openpyxl")
    st.stop()

try:
    from PIL import Image
    from streamlit_cropper import st_cropper
except ImportError:
    st.error("Missing required package(s) for the logo editor. Run: pip install pillow streamlit-cropper")
    st.stop()

# ── Constants ─────────────────────────────────────────────────────────────────
CURRENCY = "Rs."
APP_DIR = Path(__file__).resolve().parent
USERS_FILE = APP_DIR / "users.json"
APP_STATE_FILE = APP_DIR / "app_state.json"
AUDIT_FILE = APP_DIR / "audit_log.json"

DEFAULT_APP_SETTINGS = {
    "school_name": "Lahore Grammar School Accounts Office",
    "currency": "Rs.",
    "default_tone": "polite",
    "use_urdu": False,
    "duplicate_mode": "merge",
    "default_due_day": 15,
    "logo_data": "",
    "custom_aliases": {},
}

DEFAULT_CUSTOM_TEMPLATES = {
    "polite_wa":  "Hi, just a reminder that {student} has a pending fee of {amount}. Please clear it at your earliest. Thanks! - {school}",
    "firm_wa":    "Reminder: Fee of {amount} for {student} is outstanding. Please clear it soon. - {school}",
    "final_wa":   "URGENT: Final warning for {student}'s fee of {amount}. Immediate payment required. - {school}",
    "polite_letter": "Dear {parent},\n\nOur records show that the fee for {student} ({class}) has a pending balance of {amount}.\n\nPlease ignore this reminder if payment has already been made.\n\nRegards,\n{school}",
    "firm_letter":   "Dear {parent},\n\nThis is a second reminder regarding the outstanding fee for {student} ({class}) amounting to {amount}.\n\nPrompt payment is requested to avoid further reminders.\n\nRegards,\n{school}",
    "final_letter":  "Dear {parent},\n\nThis is a FINAL WARNING. The outstanding fee for {student} ({class}) of {amount} is long overdue.\n\nImmediate action is required.\n\nRegards,\n{school}",
}

FOLLOW_UP_STATUSES = ["Not Contacted", "Message Sent", "Promised", "Paid", "Disputed"]

ALIASES = {
    "student": ["student", "student name", "name", "child", "child name", "pupil"],
    "parent":  ["parent", "parent name", "father", "father name", "guardian", "guardian name"],
    "class":   ["class", "grade", "section", "class/section", "group"],
    "phone":   ["phone", "mobile", "contact", "contact no", "phone no", "cell"],
    "email":   ["email", "email address", "parent email"],
    "month":   ["month", "fee month", "billing month", "period"],
    "total":   ["total", "total fee", "fee", "amount due", "due", "payable", "invoice amount"],
    "paid":    ["paid", "paid amount", "amount paid", "received", "collection", "collected"],
    "pending": ["pending", "balance", "arrears", "outstanding", "remaining", "dues", "unpaid"],
    "status":  ["status", "payment status", "paid status"],
    "due_date":["due date", "last date", "deadline"],
}

# ── Auth helpers ──────────────────────────────────────────────────────────────
def hash_password(password: str) -> str:
    return bcrypt.hashpw(
        password.encode("utf-8"),
        bcrypt.gensalt(rounds=12)
    ).decode("utf-8")


def verify_password(password: str, stored_hash: str) -> bool:
    try:
        return bcrypt.checkpw(
            password.encode("utf-8"),
            stored_hash.encode("utf-8")
        )
    except Exception:
        return False

def load_users() -> dict:
    if not USERS_FILE.exists():
        users = {"umer": {"hash": hash_password("umer#2A"), "role": "admin"}}
        USERS_FILE.write_text(json.dumps(users), encoding="utf-8")
        return users
    try:
        
        data = json.loads(USERS_FILE.read_text(encoding="utf-8"))
        upgraded = {}
        for u, val in data.items():
            if isinstance(val, str):
                upgraded[u] = {"hash": val, "role": "admin" if u == "umer" else "user"}
            elif isinstance(val, dict):
                upgraded[u] = {"hash": val.get("hash", ""), "role": val.get("role", "user")}
            else:
                upgraded[u] = {"hash": "", "role": "user"}
        return upgraded
    except Exception:
        users = {"umer": {"hash": hash_password("umer#2A"), "role": "admin"}}
        USERS_FILE.write_text(json.dumps(users), encoding="utf-8")
        return users

def save_users(users: dict):
    USERS_FILE.write_text(json.dumps(users), encoding="utf-8")

def load_json_file(path: Path, default):
    if not path.exists():
        return default
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return default

def save_json_file(path: Path, data):
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

def load_persistent_state() -> dict:
    state = load_json_file(APP_STATE_FILE, {})
    state.setdefault("app_settings", DEFAULT_APP_SETTINGS.copy())
    state.setdefault("custom_templates", DEFAULT_CUSTOM_TEMPLATES.copy())
    state.setdefault("fee_calendar", [])
    state.setdefault("followups", {})
    return state

def save_persistent_state():
    state = {
        "app_settings": st.session_state.get("app_settings", DEFAULT_APP_SETTINGS.copy()),
        "custom_templates": st.session_state.get("custom_templates", DEFAULT_CUSTOM_TEMPLATES.copy()),
        "fee_calendar": st.session_state.get("fee_calendar", []),
        "followups": st.session_state.get("followups", {}),
    }
    save_json_file(APP_STATE_FILE, state)

def log_audit(action: str, detail: str = ""):
    log = load_json_file(AUDIT_FILE, [])
    log.append({
        "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "user": st.session_state.get("username", ""),
        "role": st.session_state.get("role", ""),
        "action": action,
        "detail": detail,
    })
    save_json_file(AUDIT_FILE, log[-500:])

def build_backup_bytes() -> bytes:
    data = load_persistent_state()
    data["users"] = load_users()
    data["audit_log"] = load_json_file(AUDIT_FILE, [])
    data["exported_at"] = datetime.now().isoformat(timespec="seconds")
    return json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8")

def restore_backup(data: dict):
    if isinstance(data.get("app_settings"), dict):
        st.session_state.app_settings = {**DEFAULT_APP_SETTINGS, **data["app_settings"]}
    if isinstance(data.get("custom_templates"), dict):
        st.session_state.custom_templates = {**DEFAULT_CUSTOM_TEMPLATES, **data["custom_templates"]}
    if isinstance(data.get("fee_calendar"), list):
        st.session_state.fee_calendar = data["fee_calendar"]
    if isinstance(data.get("followups"), dict):
        st.session_state.followups = data["followups"]
    if isinstance(data.get("followups"), dict):
        st.session_state.followups = data["followups"]
    if isinstance(data.get("users"), dict) and st.session_state.get("role") == "admin":
        save_users(data["users"])
    save_persistent_state()
    log_audit("Restore backup", "Restored app data")

# ── Data helpers ──────────────────────────────────────────────────────────────
def norm(value) -> str:
    text = "" if value is None else str(value)
    return re.sub(r"[\s_\-./()]+", " ", text.strip().lower()).strip()

def money(value) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = re.sub(r"[^0-9.\-]", "", str(value))
    if cleaned in {"", ".", "-", "-."}:
        return 0.0
    try:
        return float(cleaned)
    except ValueError:
        return 0.0

def current_currency() -> str:
    return st.session_state.get("app_settings", {}).get("currency", CURRENCY)

def ensure_app_settings() -> dict:
    cfg = DEFAULT_APP_SETTINGS.copy()
    cfg.update(st.session_state.get("app_settings") or {})
    if cfg.get("currency") not in ["Rs.", "PKR", "$", "£", "€", "AED"]:
        cfg["currency"] = CURRENCY
    if cfg.get("default_tone") not in ["polite", "firm", "final"]:
        cfg["default_tone"] = "polite"
    if cfg.get("duplicate_mode") not in ["merge", "separate"]:
        cfg["duplicate_mode"] = "merge"
    st.session_state.app_settings = cfg
    return cfg

def sync_active_report_settings(cfg: dict | None = None):
    cfg = cfg or ensure_app_settings()
    active = st.session_state.get("settings") if isinstance(st.session_state.get("settings"), dict) else {}
    active.update({
        "school_name": cfg["school_name"],
        "tone": cfg["default_tone"],
        "use_urdu": cfg["use_urdu"],
        "currency": cfg["currency"],
    })
    st.session_state.settings = active

def ensure_custom_templates() -> dict:
    templates = DEFAULT_CUSTOM_TEMPLATES.copy()
    templates.update(st.session_state.get("custom_templates") or {})
    st.session_state.custom_templates = templates
    return templates

class _SafeTemplateVars(dict):
    def __missing__(self, key):
        return "{" + key + "}"

def template_vars_for(record, school_name, currency=None) -> dict:
    return _SafeTemplateVars({
        "student": record.get("student", ""),
        "parent": record.get("parent", ""),
        "amount": fmt_amount(record.get("pending", 0), currency),
        "class": record.get("class", ""),
        "school": school_name,
        "month": record.get("month", ""),
        "due_date": record.get("due_date", ""),
    })

def render_template_text(template: str, values: dict) -> str:
    try:
        return template.format_map(values)
    except Exception:
        return template

def fmt_amount(value: float, currency: str | None = None) -> str:
    symbol = currency or current_currency()
    return f"{symbol} {value:,.0f}" if value == int(value) else f"{symbol} {value:,.2f}"

def current_date_label() -> str:
    return datetime.now().strftime("%A, %d %B %Y")

def record_key(record: dict) -> str:
    base = "|".join([
        str(record.get("student", "")).strip().lower(),
        str(record.get("parent", "")).strip().lower(),
        str(record.get("class", "")).strip().lower(),
        str(record.get("month", "")).strip().lower(),
    ])
    return hashlib.sha1(base.encode("utf-8")).hexdigest()[:16]

def valid_phone(phone: str) -> bool:
    digits = re.sub(r"[^0-9]", "", str(phone or ""))
    if digits.startswith("00"):
        digits = digits[2:]
    if digits.startswith("0"):
        return len(digits) == 11
    return len(digits) >= 10

def due_days_overdue(value: str) -> int | None:
    if not value:
        return None
    text = str(value).strip()
    formats = ["%d %B %Y", "%d %b %Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"]
    for fmt in formats:
        try:
            return (datetime.now().date() - datetime.strptime(text, fmt).date()).days
        except ValueError:
            pass
    return None

def overdue_bucket(record: dict) -> str:
    days = due_days_overdue(record.get("due_date", ""))
    if days is None:
        return "No due date"
    if days <= 0:
        return "Not overdue"
    if days <= 7:
        return "1-7 days"
    if days <= 15:
        return "8-15 days"
    if days <= 30:
        return "16-30 days"
    return "30+ days"

def recalc_analysis(analysis: dict):
    records = analysis.get("records", [])
    for r in records:
        r["total"] = money(r.get("total", 0))
        r["paid"] = money(r.get("paid", 0))
        r["pending"] = max(money(r.get("pending", 0)), 0)
        r["status"] = "Pending" if r["pending"] > 0 else "Paid"
        r["is_defaulter"] = r["pending"] > 0
    defaulters = [r for r in records if r.get("is_defaulter")]
    analysis["defaulters"] = defaulters
    analysis["total_pending"] = sum(r["pending"] for r in defaulters)
    analysis["total_expected"] = sum(r["total"] for r in records)
    analysis["total_collected"] = sum(r["paid"] for r in records)
    analysis["student_count"] = len(records)
    return analysis

def alert_counts(analysis: dict) -> dict:
    records = analysis.get("records", []) if analysis else []
    seen, dupes = set(), set()
    missing_phone = invalid_phone = overdue_30 = 0
    for r in records:
        key = (str(r.get("student", "")).lower(), str(r.get("parent", "")).lower(), str(r.get("class", "")).lower())
        if key in seen:
            dupes.add(key)
        seen.add(key)
        phone = str(r.get("phone", "")).strip()
        if not phone:
            missing_phone += 1
        elif not valid_phone(phone):
            invalid_phone += 1
        days = due_days_overdue(r.get("due_date", ""))
        if days and days > 30 and r.get("pending", 0) > 0:
            overdue_30 += 1
    return {"duplicates": len(dupes), "missing_phone": missing_phone, "invalid_phone": invalid_phone, "overdue_30": overdue_30}

def find_header_row(ws):
    best_row, best_score = 1, -1
    for row in range(1, min(ws.max_row, 15) + 1):
        values = [norm(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)]
        score = sum(1 for aliases in ALIASES.values() if any(v in aliases for v in values))
        if score > best_score:
            best_row, best_score = row, score
    return best_row

def map_headers(headers):
    normalized = [norm(h) for h in headers]
    mapping = {}
    custom_aliases = st.session_state.get("app_settings", {}).get("custom_aliases", {})
    all_aliases = {k: list(v) for k, v in ALIASES.items()}
    if isinstance(custom_aliases, dict):
        for key, values in custom_aliases.items():
            if key in all_aliases and isinstance(values, list):
                all_aliases[key].extend(norm(v) for v in values if str(v).strip())
    for key, aliases in all_aliases.items():
        for idx, header in enumerate(normalized):
            if header in aliases or any(alias in header for alias in aliases if len(alias) > 4):
                mapping[key] = idx
                break
    return mapping

def get_val(row, mapping, key, default=""):
    idx = mapping.get(key)
    if idx is None or idx >= len(row):
        return default
    v = row[idx]
    if v is None:
        return default
    return str(v).strip() if not isinstance(v, (int, float)) else v

def analyze_workbook(file_bytes: bytes, month_override=""):
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    records = []
    for ws in wb.worksheets:
        all_rows = list(ws.rows)
        if not all_rows:
            continue
        best_idx, best_score = 0, -1
        for ri in range(min(15, len(all_rows))):
            vals = [norm(c.value) for c in all_rows[ri]]
            score = sum(1 for al in ALIASES.values() if any(v in al for v in vals))
            if score > best_score:
                best_idx, best_score = ri, score
        headers = [c.value for c in all_rows[best_idx]]
        mapping = map_headers(headers)
        if "student" not in mapping:
            continue
        for row_cells in all_rows[best_idx + 1:]:
            row = [c.value for c in row_cells]
            if not any(v not in (None, "") for v in row):
                continue
            total   = money(get_val(row, mapping, "total", 0))
            paid    = money(get_val(row, mapping, "paid", 0))
            pending = money(get_val(row, mapping, "pending", 0)) if "pending" in mapping else max(total - paid, 0)
            status_norm = norm(get_val(row, mapping, "status", ""))
            is_def = pending > 0 or status_norm in {"unpaid","not paid","partial","partially paid","pending","default"}
            if status_norm in {"paid","cleared","received"} and pending < 0.01:
                is_def = False
            student = str(get_val(row, mapping, "student", "")).strip()
            if not student or norm(student) in ALIASES["student"]:
                continue
            records.append({
                "student":  student,
                "parent":   str(get_val(row, mapping, "parent", "Parent/Guardian")).strip() or "Parent/Guardian",
                "class":    str(get_val(row, mapping, "class", "")).strip(),
                "phone":    str(get_val(row, mapping, "phone", "")).strip(),
                "email":    str(get_val(row, mapping, "email", "")).strip(),
                "month":    month_override or str(get_val(row, mapping, "month", "")).strip(),
                "total":    total,
                "paid":     paid,
                "pending":  pending,
                "due_date": str(get_val(row, mapping, "due_date", "")).strip(),
                "status":   "Pending" if is_def else "Paid",
                "sheet":    ws.title,
                "is_defaulter": is_def,
            })
    wb.close()
    return records

def analyze_files(files_data, month_name="", duplicate_mode="merge"):
    all_raw = []
    for filename, data in files_data:
        recs = analyze_workbook(data, month_name)
        for r in recs:
            r["filename"] = filename
        all_raw.extend(recs)

    warnings, duplicate_audit = [], []
    occurrences = {}
    for r in all_raw:
        key = (r["student"].lower(), r["parent"].lower(), r["class"].lower())
        occurrences.setdefault(key, []).append(r)

    for key, refs in occurrences.items():
        if len(refs) > 1:
            locs = [f"{r['filename']} ({r['sheet']})" for r in refs]
            if len(set(locs)) > 1:
                warnings.append(f"Duplicate '{refs[0]['student']}' found in: {', '.join(locs)}.")

    if duplicate_mode == "separate":
        final = all_raw
    else:
        aggregated = {}
        for r in all_raw:
            key = (r["student"].lower(), r["parent"].lower(), r["class"].lower())
            if key not in aggregated:
                agg = r.copy()
                agg["months_unpaid"] = [r["month"]] if r["pending"] > 0 else []
                agg["all_months"]    = [r["month"]] if r["month"] else []
                aggregated[key] = agg
            else:
                agg = aggregated[key]
                agg["total"] += r["total"]; agg["paid"] += r["paid"]; agg["pending"] += r["pending"]
                if r["pending"] > 0 and r["month"] and r["month"] not in agg["months_unpaid"]:
                    agg["months_unpaid"].append(r["month"])
                if r["month"] and r["month"] not in agg["all_months"]:
                    agg["all_months"].append(r["month"])
                if not agg["phone"]: agg["phone"] = r["phone"]
                if not agg["email"]: agg["email"] = r["email"]
        for agg in aggregated.values():
            agg["month"] = ", ".join(agg["months_unpaid"]) if agg["months_unpaid"] else ", ".join(sorted(agg["all_months"]))
            agg["status"] = "Pending" if agg["pending"] > 0 else "Paid"
            agg["is_defaulter"] = agg["pending"] > 0
        final = list(aggregated.values())

    defaulters = [r for r in final if r["pending"] > 0]
    return {
        "records":        final,
        "defaulters":     defaulters,
        "total_pending":  sum(r["pending"] for r in defaulters),
        "total_expected": sum(r["total"]   for r in final),
        "total_collected":sum(r["paid"]    for r in final),
        "student_count":  len(final),
        "warnings":       warnings,
        "duplicate_audit":duplicate_audit,
        "month_override": month_name,
        "duplicate_mode": duplicate_mode,
    }

def letter_for(record, tone="polite", school_name="Accounts Office", use_urdu=False, currency=None):
    amount  = fmt_amount(record["pending"], currency)
    student = record["student"]
    parent  = record["parent"]
    cls     = f" (Class: {record['class']})" if record["class"] else ""
    if use_urdu:
        if tone == "firm":
            msg = f"محترم والدین، یہ {student}{cls} کی بقایا فیس {amount} کے بارے میں دوسری یاد دہانی ہے۔ براہ کرم فوری ادائیگی کو یقینی بنائیں۔"
        elif tone == "final":
            msg = f"محترم والدین، یہ فیس کی ادائیگی کے لیے آخری وارننگ ہے۔ {student}{cls} کی فیس {amount} کافی عرصہ سے واجب الادا ہے۔"
        else:
            msg = f"محترم والدین، ہمارے ریکارڈ کے مطابق {student}{cls} کی فیس کے {amount} واجب الادا ہیں۔"
        return f"{parent} کے نام،\n\n{msg}\n\nشکریہ،\n{school_name}"
    template = ensure_custom_templates().get(f"{tone}_letter")
    if template:
        return render_template_text(template, template_vars_for(record, school_name, currency))
    month    = f" for {record['month']}" if record["month"] else ""
    due_date = f" Kindly clear the dues by {record['due_date']}." if record["due_date"] else ""
    if tone == "firm":
        intro   = f"This is a second reminder regarding the outstanding fee{month} for {student}{cls}."
        closing = "Prompt payment is requested to avoid further reminders."
    elif tone == "final":
        intro   = f"This is a FINAL WARNING. The outstanding fee{month} for {student}{cls} is long overdue."
        closing = "Immediate action is required to avoid potential suspension of services."
    else:
        intro   = f"Our records show that the fee{month} for {student}{cls} has a pending balance of {amount}.{due_date}"
        closing = "Please ignore this reminder if payment has already been made. For any correction, kindly contact us."
    return f"Dear {parent},\n\n{intro}\n\n{closing}\n\nRegards,\n{school_name}"

def whatsapp_msg_for(record, tone="polite", school_name="Accounts Office", use_urdu=False, currency=None):
    amount = fmt_amount(record["pending"], currency)
    student = record["student"]
    if use_urdu:
        if tone == "firm":   return f"یاد دہانی: {student} کی فیس {amount} واجب الادا ہے۔ - {school_name}"
        elif tone == "final": return f"آخری انتباہ: {student} کی فیس {amount} واجب الادا ہے۔ - {school_name}"
        return f"سلام: {student} کی فیس {amount} واجب الادا ہے۔ - {school_name}"
    template = ensure_custom_templates().get(f"{tone}_wa")
    if template:
        return render_template_text(template, template_vars_for(record, school_name, currency))
    if tone == "firm":    return f"Reminder: Fee of {amount} for {student} is outstanding. Please clear it soon. - {school_name}"
    elif tone == "final": return f"URGENT: Final warning for {student}'s fee of {amount}. Immediate payment required. - {school_name}"
    return f"Hi, just a reminder that {student} has a pending fee of {amount}. Please clear it at your earliest. Thanks! - {school_name}"

def wa_click_link(phone: str, message: str) -> str:
    phone = re.sub(r"[^0-9]", "", phone)
    if not phone:
        return ""
    if phone.startswith("00"):   phone = phone[2:]
    if phone.startswith("0"):    phone = "92" + phone[1:11]
    return f"https://wa.me/{phone}?text={quote(message)}"

def build_sample_workbook_bytes():
    wb = Workbook(); ws = wb.active; ws.title = "June Fees"
    ws.append(["Student","Parent/Guardian","Class","Phone","Email","Month","Total Fee","Paid","Pending","Due Date","Status"])
    ws.append(["Ali Khan","Mr. Khan","Grade 7","03001234567","ali@example.com","June 2026",15000,5000,10000,"15 June 2026","Pending"])
    ws.append(["Sara Ahmed","Mrs. Ahmed","Grade 8","03007654321","sara@example.com","June 2026",12000,12000,0,"15 June 2026","Paid"])
    ws.append(["Hassan Raza","Mr. Raza","Grade 7","03001112223","hassan@example.com","June 2026",15000,0,15000,"15 June 2026","Unpaid"])
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

def write_report_bytes(analysis, school_name, tone, use_urdu, currency=None) -> bytes:
    wb = Workbook()
    summary = wb.active; summary.title = "Summary Dashboard"
    summary.append(["School Name", school_name]); summary.append([])
    summary.append(["Key Metric", "Value"])
    total_s = analysis["student_count"]
    def_cnt = len(analysis["defaulters"])
    t_exp   = analysis["total_expected"]
    t_col   = analysis["total_collected"]
    t_pend  = analysis["total_pending"]
    crate   = (t_col / t_exp * 100) if t_exp > 0 else 0
    for row in [["Total Students",total_s],["Defaulters",def_cnt],["Total Expected",t_exp],
                ["Total Collected",t_col],["Total Pending",t_pend],["Collection Rate",f"{crate:.1f}%"]]:
        summary.append(row)
    summary.append([]); summary.append(["Class-wise Breakdown","Pending Amount"])
    class_stats = {}
    for r in analysis["defaulters"]:
        c = r["class"] or "Uncategorized"
        class_stats[c] = class_stats.get(c, 0) + r["pending"]
    for c, amt in sorted(class_stats.items()):
        summary.append([c, amt])

    ws = wb.create_sheet("Defaulters")
    ws.append(["Student","Parent/Guardian","Class","Phone","Email","Month","Total Fee","Paid","Pending","Due Date","Letter Draft","WhatsApp Link"])
    for rec in analysis["defaulters"]:
        msg  = whatsapp_msg_for(rec, tone, school_name, use_urdu, currency)
        link = wa_click_link(rec.get("phone",""), msg)
        ws.append([rec["student"],rec["parent"],rec["class"],rec["phone"],rec["email"],
                   rec["month"],rec["total"],rec["paid"],rec["pending"],rec["due_date"],
                   letter_for(rec, tone, school_name, use_urdu, currency), link])
        if link:
            cell = ws.cell(ws.max_row, ws.max_column)
            cell.hyperlink = link; cell.font = Font(color="0563C1", underline="single")

    fill = PatternFill("solid", fgColor="1F4E79"); hfont = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.fill = fill; cell.font = hfont; cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = Border(bottom=thin); cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col in range(1, sheet.max_column + 1):
            width = min(max(12, max(len(str(sheet.cell(r, col).value or "")) for r in range(1, sheet.max_row + 1)) + 2), 55)
            sheet.column_dimensions[get_column_letter(col)].width = width
        sheet.freeze_panes = "A2"; sheet.auto_filter.ref = sheet.dimensions

    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

def write_letters_bytes(analysis, school_name, tone, use_urdu, currency=None):
    direction = "rtl" if use_urdu else "ltr"
    cards = []
    logo_data = st.session_state.get("app_settings", {}).get("logo_data", "")
    logo_html = f"<img src='{logo_data}' style='max-height:70px;margin-bottom:18px;' alt='School logo'>" if logo_data else ""
    for rec in analysis["defaulters"]:
        import html as html_mod
        body = html_mod.escape(letter_for(rec, tone, school_name, use_urdu, currency)).replace("\n","<br>")
        msg  = whatsapp_msg_for(rec, tone, school_name, use_urdu, currency)
        link = wa_click_link(rec.get("phone",""), msg)
        link_html = f"<div style='margin-top:20px;'><a href='{html_mod.escape(link)}' style='color:#25D366;font-weight:bold;' class='no-print'>Send via WhatsApp</a></div>" if link else ""
        cards.append(f"<section class='letter' dir='{direction}'>{logo_html}{body}{link_html}</section>")
    html_out = (
        "<!doctype html><html><head><meta charset='utf-8'><title>Reminder Letters</title>"
        f"<style>body{{font-family:Arial,sans-serif;margin:32px;direction:{direction}}} "
        ".letter{page-break-after:always;max-width:760px;margin:0 auto 36px;padding:28px;border:1px solid #d8dee8;border-radius:8px;line-height:1.55}"
        "@media print{.letter{border:0;margin:0;padding:0}.no-print{display:none}}</style></head>"
        f"<body>{''.join(cards) if cards else '<p>No pending fee records found.</p>'}</body></html>"
    )
    return html_out.encode("utf-8")

def write_pdf_letters_bytes(analysis, school_name, tone, use_urdu, currency=None):
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import inch
    from reportlab.lib import colors

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4

    for rec in analysis["defaulters"] or [{}]:
        # Professional Layout
        c.setFont("Helvetica-Bold", 18)
        c.drawString(1 * inch, height - 1 * inch, school_name[:50])
        
        c.setStrokeColor(colors.lightgrey)
        c.line(1 * inch, height - 1.2 * inch, width - 1 * inch, height - 1.2 * inch)

        c.setFont("Helvetica", 11)
        y = height - 1.6 * inch
        
        letter_content = letter_for(rec, tone, school_name, use_urdu, currency) if rec else "No pending fee records found."
        
        # Simple text wrapping for PDF
        import textwrap
        lines = []
        for p in letter_content.splitlines():
            if not p: lines.append("")
            else: lines.extend(textwrap.wrap(p, width=90))
            
        for line in lines:
            c.drawString(1 * inch, y, line)
            y -= 14
            if y < 1 * inch:
                c.showPage()
                c.setFont("Helvetica", 11)
                y = height - 1 * inch
                
        c.showPage()
    c.save()
    return buf.getvalue()

def write_wa_messages_bytes(analysis, school_name, tone, use_urdu, currency=None):
    parts = []
    for rec in analysis["defaulters"]:
        msg  = whatsapp_msg_for(rec, tone, school_name, use_urdu, currency)
        link = wa_click_link(rec.get("phone",""), msg)
        parts.append(f"Student: {rec['student']}\nParent: {rec['parent']}\nPhone: {rec.get('phone','')}\nMessage: {msg}\nWhatsApp Link: {link}")
    return "\n\n---\n\n".join(parts).encode("utf-8")

def write_csv_bytes(analysis, school_name, tone, use_urdu, currency=None):
    buf = io.StringIO()
    fields = ["student","parent","class","phone","email","month","total","paid","pending","due_date","whatsapp_link"]
    writer = csv.DictWriter(buf, fieldnames=fields)
    writer.writeheader()
    for rec in analysis["defaulters"]:
        row = {k: rec.get(k,"") for k in fields}
        msg = whatsapp_msg_for(rec, tone, school_name, use_urdu, currency)
        row["whatsapp_link"] = wa_click_link(rec.get("phone",""), msg)
        writer.writerow(row)
    return buf.getvalue().encode("utf-8")

def build_zip(files: dict[str, bytes]) -> bytes:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for name, data in files.items():
            zf.writestr(name, data)
    return buf.getvalue()

def class_wise_zip_bytes(analysis, school_name, tone, use_urdu, currency=None) -> bytes:
    files = {}
    classes = sorted({r.get("class") or "Uncategorized" for r in analysis.get("defaulters", [])})
    for cls in classes:
        safe = re.sub(r"[^A-Za-z0-9_-]+", "_", cls).strip("_") or "Uncategorized"
        subset = dict(analysis)
        subset["records"] = [r for r in analysis.get("records", []) if (r.get("class") or "Uncategorized") == cls]
        subset["defaulters"] = [r for r in analysis.get("defaulters", []) if (r.get("class") or "Uncategorized") == cls]
        recalc_analysis(subset)
        files[f"{safe}/fee_report_{safe}.xlsx"] = write_report_bytes(subset, school_name, tone, use_urdu, currency)
        files[f"{safe}/letters_{safe}.pdf"] = write_pdf_letters_bytes(subset, school_name, tone, use_urdu, currency)
    if not files:
        files["no_defaulters.txt"] = b"No pending fee records found."
    return build_zip(files)

# ── Streamlit UI ──────────────────────────────────────────────────────────────
# (st.set_page_config moved to top)

# ── Design system ─────────────────────────────────────────────────────────────
THEME_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

html, body, [class*="css"] {
    font-family: 'Inter', sans-serif;
}

/* ── Page background ── */
.stApp {
    background: #F7F9FC;
}

/* ── Hide default Streamlit chrome ── */
#MainMenu, footer, header { visibility: hidden; }
.block-container { padding-top: 1.5rem !important; padding-bottom: 3rem !important; }

/* ── Sidebar ── */
[data-testid="stSidebar"] {
    background: linear-gradient(160deg, #0D1B2A 0%, #1B3A5C 100%) !important;
    border-right: none !important;
}
[data-testid="stSidebar"] > div:first-child { padding: 0 !important; }

.sidebar-brand {
    padding: 28px 24px 20px;
    border-bottom: 1px solid rgba(255,255,255,0.08);
    margin-bottom: 8px;
}
.sidebar-brand .app-icon {
    width: 42px; height: 42px;
    background: linear-gradient(135deg, #3B82F6, #06B6D4);
    border-radius: 10px;
    display: flex; align-items: center; justify-content: center;
    font-size: 20px; margin-bottom: 12px;
}
.sidebar-brand .app-name {
    font-size: 0.92rem; font-weight: 700; color: #F0F6FF;
    line-height: 1.3; letter-spacing: -0.01em;
}
.sidebar-brand .app-sub {
    font-size: 0.72rem; color: #64748B; margin-top: 2px; font-weight: 400;
}

.sidebar-user {
    padding: 14px 24px;
    margin: 0 12px 4px;
    background: rgba(255,255,255,0.05);
    border-radius: 10px;
    display: flex; align-items: center; gap: 10px;
}
.sidebar-user .avatar {
    width: 34px; height: 34px; border-radius: 50%;
    background: linear-gradient(135deg, #3B82F6, #8B5CF6);
    display: flex; align-items: center; justify-content: center;
    font-size: 14px; font-weight: 700; color: white; flex-shrink: 0;
}
.sidebar-user .user-name { font-size: 0.85rem; font-weight: 600; color: #E2E8F0; }
.sidebar-user .user-role {
    font-size: 0.68rem; font-weight: 500; color: #3B82F6;
    background: rgba(59,130,246,0.15); border-radius: 4px;
    padding: 1px 6px; display: inline-block; margin-top: 2px;
    text-transform: uppercase; letter-spacing: 0.05em;
}
.sidebar-date {
    margin: 0 12px 16px;
    padding: 10px 14px;
    border: 1px solid rgba(148,163,184,0.16);
    border-radius: 8px;
    background: rgba(15,23,42,0.42);
}
.sidebar-date .date-label {
    font-size: 0.64rem;
    color: #64748B;
    text-transform: uppercase;
    letter-spacing: 0.08em;
    font-weight: 700;
}
.sidebar-date .date-value {
    margin-top: 2px;
    color: #E2E8F0;
    font-size: 0.78rem;
    font-weight: 600;
    line-height: 1.35;
}

/* Nav radio */
[data-testid="stSidebar"] .stRadio > label { display: none !important; }
[data-testid="stSidebar"] .stRadio > div {
    gap: 2px !important;
    padding: 8px 12px;
}
[data-testid="stSidebar"] .stRadio > div > label {
    display: flex !important; align-items: center !important;
    padding: 10px 14px !important; border-radius: 8px !important;
    cursor: pointer !important; transition: all 0.15s !important;
    color: #FFFFFF !important; font-size: 0.875rem !important;
    font-weight: 500 !important; border: none !important;
}
[data-testid="stSidebar"] .stRadio > div > label p,
[data-testid="stSidebar"] .stRadio > div > label span {
    color: #FFFFFF !important;
}
[data-testid="stSidebar"] .stRadio > div > label:hover {
    background: rgba(255,255,255,0.08) !important;
    color: #FFFFFF !important;
}
[data-testid="stSidebar"] .stRadio > div > label:hover p,
[data-testid="stSidebar"] .stRadio > div > label:hover span {
    color: #FFFFFF !important;
}
[data-testid="stSidebar"] .stRadio [data-checked="true"] > label,
[data-testid="stSidebar"] .stRadio > div > label[data-checked="true"] {
    background: rgba(59,130,246,0.22) !important;
    color: #FFFFFF !important;
}
[data-testid="stSidebar"] .stRadio [data-checked="true"] > label p,
[data-testid="stSidebar"] .stRadio [data-checked="true"] > label span,
[data-testid="stSidebar"] .stRadio > div > label[data-checked="true"] p,
[data-testid="stSidebar"] .stRadio > div > label[data-checked="true"] span {
    color: #FFFFFF !important;
}
[data-testid="stSidebar"] .stRadio > div > label > div:first-child { display: none !important; }

/* Sidebar buttons */
[data-testid="stSidebar"] .stButton > button {
    background: rgba(255,255,255,0.06) !important;
    color: #94A3B8 !important; border: 1px solid rgba(255,255,255,0.08) !important;
    border-radius: 8px !important; font-size: 0.83rem !important;
    font-weight: 500 !important; padding: 9px 14px !important;
    transition: all 0.15s !important;
}
[data-testid="stSidebar"] .stButton > button:hover {
    background: rgba(255,255,255,0.1) !important;
    color: #E2E8F0 !important; border-color: rgba(255,255,255,0.15) !important;
}
.sidebar-divider {
    height: 1px; background: rgba(255,255,255,0.07);
    margin: 10px 24px;
}

/* ── Page header ── */
.page-header {
    background: linear-gradient(135deg, #0D1B2A 0%, #1B3A5C 60%, #1E4D8C 100%);
    border-radius: 16px;
    padding: 28px 32px;
    margin-bottom: 24px;
    position: relative;
    overflow: hidden;
}
.page-header::after {
    content: '';
    position: absolute; top: -40px; right: -40px;
    width: 200px; height: 200px;
    background: radial-gradient(circle, rgba(59,130,246,0.15) 0%, transparent 70%);
    border-radius: 50%;
}
.page-header h1 {
    font-size: 1.6rem; font-weight: 700; color: #F0F6FF;
    margin: 0 0 6px; letter-spacing: -0.02em;
}
.page-header p { font-size: 0.875rem; color: #64748B; margin: 0; }

/* ── Cards / panels ── */
.panel {
    background: #FFFFFF;
    border: 1px solid #E8EDF5;
    border-radius: 14px;
    padding: 22px 24px;
    margin-bottom: 16px;
    box-shadow: 0 1px 4px rgba(0,0,0,0.04);
}
.panel-title {
    font-size: 0.78rem; font-weight: 700; color: #64748B;
    text-transform: uppercase; letter-spacing: 0.07em;
    margin-bottom: 16px; display: flex; align-items: center; gap: 6px;
}
.panel-title::after {
    content: ''; flex: 1; height: 1px; background: #F1F5F9;
}

/* ── Metric cards ── */
.metrics-row { display: flex; gap: 12px; margin-bottom: 20px; }
.metric-card {
    flex: 1; background: #FFFFFF;
    border: 1px solid #E8EDF5; border-radius: 12px;
    padding: 16px 20px;
    box-shadow: 0 1px 3px rgba(0,0,0,0.04);
    position: relative; overflow: hidden;
}
.metric-card::before {
    content: ''; position: absolute;
    top: 0; left: 0; right: 0; height: 3px;
    background: var(--bar-color, #3B82F6);
}
.metric-card.green::before  { background: #10B981; }
.metric-card.red::before    { background: #EF4444; }
.metric-card.amber::before  { background: #F59E0B; }
.metric-card.blue::before   { background: #3B82F6; }
.metric-card.indigo::before { background: #6366F1; }
.metric-label {
    font-size: 0.7rem; font-weight: 600; color: #94A3B8;
    text-transform: uppercase; letter-spacing: 0.07em; margin-bottom: 8px;
}
.metric-value {
    font-size: 1.55rem; font-weight: 700; color: #0F172A;
    letter-spacing: -0.02em; line-height: 1;
}
.metric-sub {
    font-size: 0.72rem; color: #94A3B8; margin-top: 6px;
}

/* ── Progress bar ── */
.progress-wrap {
    background: #F1F5F9; border-radius: 99px; height: 8px;
    overflow: hidden; margin: 4px 0 8px;
}
.progress-fill {
    height: 100%; border-radius: 99px;
    background: linear-gradient(90deg, #10B981, #3B82F6);
    transition: width 0.6s ease;
}
.progress-label {
    font-size: 0.75rem; color: #64748B; font-weight: 500;
    display: flex; justify-content: space-between; margin-bottom: 4px;
}

/* ── Warning / info banners ── */
.banner {
    border-radius: 10px; padding: 12px 16px;
    font-size: 0.85rem; margin-bottom: 10px;
    display: flex; align-items: flex-start; gap: 10px;
}
.banner.warn  { background: #FFFBEB; border: 1px solid #FDE68A; color: #92400E; }
.banner.info  { background: #EFF6FF; border: 1px solid #BFDBFE; color: #1E40AF; }
.banner.empty { background: #F0FDF4; border: 1px solid #BBF7D0; color: #14532D; padding: 24px; text-align: center; border-radius: 12px; }
.banner-icon  { font-size: 1rem; flex-shrink: 0; margin-top: 1px; }

/* ── Section divider ── */
.section-sep {
    display: flex; align-items: center; gap: 10px;
    margin: 24px 0 16px;
}
.section-sep span {
    font-size: 0.78rem; font-weight: 700; color: #475569;
    text-transform: uppercase; letter-spacing: 0.07em; white-space: nowrap;
}
.section-sep::before, .section-sep::after {
    content: ''; flex: 1; height: 1px; background: #E2E8F0;
}

/* ── Download buttons row ── */
.dl-grid { display: grid; grid-template-columns: repeat(5, 1fr); gap: 10px; margin-bottom: 20px; }
.dl-btn {
    background: #FFFFFF; border: 1px solid #E2E8F0;
    border-radius: 10px; padding: 14px 10px;
    text-align: center; cursor: pointer; transition: all 0.15s;
    text-decoration: none; color: #1E293B;
    display: flex; flex-direction: column; align-items: center; gap: 6px;
}
.dl-btn:hover { border-color: #3B82F6; box-shadow: 0 0 0 3px rgba(59,130,246,0.1); }
.dl-btn .dl-icon { font-size: 1.4rem; }
.dl-btn .dl-label { font-size: 0.72rem; font-weight: 600; color: #475569; }

/* ── Form inputs ── */
.stTextInput > label, .stSelectbox > label, .stCheckbox > label {
    font-size: 0.8rem !important; font-weight: 600 !important;
    color: #475569 !important; margin-bottom: 4px !important;
}
.stTextInput input, .stSelectbox select {
    border-radius: 8px !important; border-color: #E2E8F0 !important;
    font-size: 0.875rem !important;
}
.stTextInput input:focus, .stSelectbox select:focus {
    border-color: #3B82F6 !important;
    box-shadow: 0 0 0 3px rgba(59,130,246,0.12) !important;
}

/* ── Primary button ── */
.stButton > button[kind="primary"] {
    background: linear-gradient(135deg, #2563EB, #1D4ED8) !important;
    color: #FFFFFF !important;
    border: none !important; border-radius: 9px !important;
    font-weight: 600 !important; font-size: 0.875rem !important;
    letter-spacing: -0.01em !important;
    box-shadow: 0 2px 8px rgba(37,99,235,0.35) !important;
    transition: all 0.15s !important;
}
.stButton > button[kind="primary"] p,
.stButton > button[kind="primary"] span {
    color: #FFFFFF !important;
}
.stButton > button[kind="primary"]:hover {
    background: linear-gradient(135deg, #1D4ED8, #1E40AF) !important;
    color: #FFFFFF !important;
    box-shadow: 0 4px 14px rgba(37,99,235,0.45) !important;
    transform: translateY(-1px) !important;
}

/* ── Letter preview card ── */
.letter-card {
    background: #FAFBFF; border: 1px solid #E2E8F0;
    border-radius: 12px; padding: 24px 28px;
    font-family: Georgia, 'Times New Roman', serif;
    font-size: 0.9rem; line-height: 1.8; color: #1E293B;
    white-space: pre-wrap; margin-top: 12px;
}

/* ── Login page ── */
.login-wrap {
    max-width: 420px; margin: 60px auto;
    background: #FFFFFF; border: 1px solid #E8EDF5;
    border-radius: 20px; padding: 40px 44px;
    box-shadow: 0 8px 40px rgba(0,0,0,0.08);
}
.login-logo {
    width: 54px; height: 54px;
    background: linear-gradient(135deg, #2563EB, #06B6D4);
    border-radius: 14px; display: flex; align-items: center;
    justify-content: center; font-size: 24px;
    margin: 0 auto 20px;
}
.login-title {
    font-size: 1.4rem; font-weight: 700; color: #0F172A;
    text-align: center; margin-bottom: 4px; letter-spacing: -0.02em;
}
.login-sub {
    font-size: 0.83rem; color: #94A3B8; text-align: center; margin-bottom: 28px;
}

/* ── Impersonation banner ── */
.impersonation-banner {
    display: flex; align-items: center; justify-content: space-between;
    background: linear-gradient(90deg, #7C3AED 0%, #4F46E5 100%);
    border-radius: 10px; padding: 10px 18px;
    margin-bottom: 18px; color: #fff;
    font-size: 0.85rem; font-weight: 500;
    box-shadow: 0 2px 8px rgba(79,70,229,0.25);
}
.impersonation-banner strong { font-weight: 700; }
.impersonation-banner .imp-tag {
    background: rgba(255,255,255,0.2); border-radius: 5px;
    padding: 2px 8px; font-size: 0.75rem; font-weight: 600;
    text-transform: uppercase; letter-spacing: 0.05em; margin-right: 10px;
}

/* ── Admin panel ── */
.user-row {
    display: flex; align-items: center; gap: 12px;
    padding: 12px 16px; background: #FAFBFF;
    border: 1px solid #E8EDF5; border-radius: 10px; margin-bottom: 8px;
}
.user-row .u-avatar {
    width: 36px; height: 36px; border-radius: 50%;
    background: linear-gradient(135deg, #6366F1, #8B5CF6);
    display: flex; align-items: center; justify-content: center;
    font-size: 14px; font-weight: 700; color: white; flex-shrink: 0;
}
.user-row .u-name { font-weight: 600; color: #1E293B; font-size: 0.875rem; }
.user-row .u-role {
    font-size: 0.68rem; font-weight: 600; padding: 2px 7px;
    border-radius: 5px; text-transform: uppercase; letter-spacing: 0.05em;
}
.u-role.admin { background: #EFF6FF; color: #1D4ED8; }
.u-role.user  { background: #F0FDF4; color: #15803D; }

/* ── File uploader ── */
[data-testid="stFileUploader"] {
    border: 2px dashed #CBD5E1 !important;
    border-radius: 12px !important;
    background: #F8FAFC !important;
    transition: border-color 0.2s !important;
}
[data-testid="stFileUploader"]:hover {
    border-color: #3B82F6 !important;
    background: #EFF6FF !important;
}

/* ── Expander ── */
.streamlit-expanderHeader {
    background: #F8FAFC !important;
    border: 1px solid #E2E8F0 !important;
    border-radius: 10px !important;
    font-weight: 600 !important; font-size: 0.875rem !important;
    color: #1E293B !important;
}
.streamlit-expanderContent {
    border: 1px solid #E2E8F0 !important;
    border-top: none !important;
    border-radius: 0 0 10px 10px !important;
    background: #FAFBFF !important;
}

/* ── Table ── */
[data-testid="stDataFrame"] { border-radius: 10px; overflow: hidden; }

/* ── Download button ── */
[data-testid="stDownloadButton"] > button {
    border-radius: 9px !important; font-size: 0.8rem !important;
    font-weight: 600 !important; border: 1px solid #E2E8F0 !important;
    background: #FFFFFF !important; color: #374151 !important;
    transition: all 0.15s !important;
}
[data-testid="stDownloadButton"] > button:hover {
    border-color: #3B82F6 !important;
    color: #2563EB !important;
    box-shadow: 0 0 0 3px rgba(59,130,246,0.1) !important;
}
</style>
"""

# Cross-reload persistence
# Manual browser refreshes wipe st.session_state, so important state is
# stored server-side under the sid token in the URL. The app does not
# auto-refresh while someone is typing.
_PERSIST_TTL_SECONDS = 24 * 60 * 60  # drop stale sessions after a day

@st.cache_resource
def _get_store() -> dict:
    """One dict, shared by every browser tab and every rerun, for the life of
    the server process. st.cache_resource is required here (not a plain
    module-level global) because Streamlit re-executes this script's
    top-level code on every single rerun — a plain global would silently
    reset back to an empty dict each time, breaking persistence entirely."""
    return {}

def _persist_save():
    """Mirror the bits of session_state that must survive a full page reload
    into the shared server-side store, under this browser tab's token."""
    sid = st.session_state.get("_sid")
    if not sid:
        return
    store = _get_store()
    store[sid] = {
        "saved_at":  time.time(),
        "logged_in": st.session_state.get("logged_in", False),
        "username":  st.session_state.get("username", ""),
        "role":      st.session_state.get("role", ""),
        "analysis":  st.session_state.get("analysis"),
        "source_files": st.session_state.get("source_files"),
        "source_month": st.session_state.get("source_month", ""),
        "settings":  st.session_state.get("settings"),
        "app_settings": st.session_state.get("app_settings"),
        "custom_templates": st.session_state.get("custom_templates"),
        "fee_calendar": st.session_state.get("fee_calendar"),
        "page":      st.session_state.get("page", "report"),
        "admin_username": st.session_state.get("admin_username", ""),
        "admin_role":     st.session_state.get("admin_role", ""),
    }

def _persist_load(sid: str) -> bool:
    store = _get_store()
    now = time.time()
    for k in [k for k, v in store.items() if now - v.get("saved_at", 0) > _PERSIST_TTL_SECONDS]:
        store.pop(k, None)
    data = store.get(sid)
    if not data:
        return False
    st.session_state.logged_in = data.get("logged_in", False)
    st.session_state.username  = data.get("username", "")
    st.session_state.role      = data.get("role", "")
    st.session_state.analysis  = data.get("analysis")
    st.session_state.source_files = data.get("source_files")
    st.session_state.source_month = data.get("source_month", "")
    st.session_state.settings  = data.get("settings")
    if data.get("app_settings"):
        st.session_state.app_settings = data["app_settings"]
    if data.get("custom_templates"):
        st.session_state.custom_templates = data["custom_templates"]
    if data.get("fee_calendar") is not None:
        st.session_state.fee_calendar = data["fee_calendar"]
    st.session_state.page      = data.get("page", "report")
    st.session_state.admin_username = data.get("admin_username", "")
    st.session_state.admin_role     = data.get("admin_role", "")
    st.session_state._sid      = sid
    return True


# ── Restore session after a real browser refresh, if we have a valid token ────
if "_sid" not in st.session_state:
    try:
        qp_sid = st.query_params.get("sid")
    except Exception:
        qp_sid = (st.experimental_get_query_params().get("sid", [None]) or [None])[0]
    if not (qp_sid and _persist_load(qp_sid)):
        st.session_state._sid = secrets.token_urlsafe(24)
        try:
            st.query_params["sid"] = st.session_state._sid
        except Exception:
            st.experimental_set_query_params(sid=st.session_state._sid)


# ── Session state ─────────────────────────────────────────────────────────────
if "initialized" not in st.session_state:
    persisted = load_persistent_state()
    st.session_state.app_settings = persisted.get("app_settings", DEFAULT_APP_SETTINGS.copy())
    st.session_state.custom_templates = persisted.get("custom_templates", DEFAULT_CUSTOM_TEMPLATES.copy())
    st.session_state.fee_calendar = persisted.get("fee_calendar", [])
    st.session_state.followups = persisted.get("followups", {})
    st.session_state.message_history = persisted.get("message_history", {})
    st.session_state.initialized = True

if "logged_in" not in st.session_state:
    st.session_state.logged_in = False
    st.session_state.username  = ""
    st.session_state.role      = ""
if "analysis" not in st.session_state:
    st.session_state.analysis = None
if "source_files" not in st.session_state:
    st.session_state.source_files = None
if "source_month" not in st.session_state:
    st.session_state.source_month = ""
if "page" not in st.session_state:
    st.session_state.page = "report"
if "admin_username" not in st.session_state:
    st.session_state.admin_username = ""   # stores the real admin while impersonating
if "admin_role" not in st.session_state:
    st.session_state.admin_role = ""

# ── Login page ────────────────────────────────────────────────────────────────
def show_login():
    st.markdown(THEME_CSS, unsafe_allow_html=True)
    _, col, _ = st.columns([1, 1.1, 1])
    with col:
        st.markdown("""
        <div class="login-wrap">
            <div class="login-logo">📋</div>
            <div class="login-title">Fee Defaulter Portal</div>
            <div class="login-sub">Sign in to access your school's fee reports</div>
        </div>
        """, unsafe_allow_html=True)
        with st.form("login_form"):
            username = st.text_input("Username", placeholder="Enter your username")
            password = st.text_input("Password", type="password", placeholder="Enter your password")
            submitted = st.form_submit_button("Sign In →", use_container_width=True, type="primary")
            if submitted:
                users = load_users()
                if username in users and verify_password(password, users[username]["hash"]):
                    st.session_state.logged_in = True
                    st.session_state.username  = username
                    st.session_state.role      = users[username]["role"]
                    st.rerun()
                else:
                    st.error("Incorrect username or password. Please try again.")

# ── Sidebar ───────────────────────────────────────────────────────────────────
def render_sidebar():
    with st.sidebar:
        uname  = st.session_state.username
        initials = uname[:2].upper() if uname else "?"
        today_label = current_date_label()
        logo_data = st.session_state.get("app_settings", {}).get("logo_data", "")
        if logo_data:
            brand_icon_html = f'<img src="{logo_data}" alt="School logo" style="width:100%;height:100%;object-fit:cover;border-radius:10px;">'
        else:
            brand_icon_html = "📋"
        st.markdown(f"""
        <div class="sidebar-brand">
            <div class="app-icon">{brand_icon_html}</div>
            <div class="app-name">Fee Defaulter Report Generator</div>
            <div class="app-sub">School Finance Tool</div>
        </div>
        <div class="sidebar-user" style="margin-bottom:16px;">
            <div class="avatar">{initials}</div>
            <div>
                <div class="user-name">{uname}</div>
                <span class="user-role">{st.session_state.role}</span>
            </div>
        </div>
        <div class="sidebar-date">
            <div class="date-label">Today</div>
            <div class="date-value">{today_label}</div>
        </div>
        """, unsafe_allow_html=True)

        nav_options = [
            "📊  Generate Report",
            "📈  Analytics",
            "👥  Student Directory",
            "✉️  Message Templates",
            "📅  Fee Calendar",
            "🖨️  Export Center",
            "⚙️  Settings",
            "ℹ️  Help",
        ]
        
        # Map internal page keys to radio index for persistence after reload
        curr = st.session_state.get("page", "report")
        idx = 0
        if curr == "analytics": idx = 1
        elif curr == "students": idx = 2
        elif curr == "templates": idx = 3
        elif curr == "calendar": idx = 4
        elif curr == "export": idx = 5
        elif curr == "settings": idx = 6
        elif curr == "help": idx = 7
        
        page = st.radio("Navigation", nav_options, index=idx, label_visibility="collapsed")
        
        # Sync session_state.page with radio selection
        if "Analytics" in page: st.session_state.page = "analytics"
        elif "Student Directory" in page: st.session_state.page = "students"
        elif "Message Templates" in page: st.session_state.page = "templates"
        elif "Fee Calendar" in page: st.session_state.page = "calendar"
        elif "Export Center" in page: st.session_state.page = "export"
        elif "Settings" in page: st.session_state.page = "settings"
        elif "Help" in page: st.session_state.page = "help"
        elif st.session_state.page != "admin": # Don't overwrite admin if we just switched
            st.session_state.page = "report"

        st.markdown('<div class="sidebar-divider"></div>', unsafe_allow_html=True)

        if st.session_state.role == "admin":
            if st.button("🔐  Admin Panel", use_container_width=True):
                st.session_state.page = "admin"
                st.rerun()

        if st.button("🚪  Sign Out", use_container_width=True):
            _get_store().pop(st.session_state.get("_sid", ""), None)
            for k in ["logged_in","username","role","analysis","source_files","source_month","page","settings","admin_username","admin_role"]:
                st.session_state.pop(k, None)
            st.session_state._sid = secrets.token_urlsafe(24)
            try:
                st.query_params["sid"] = st.session_state._sid
            except Exception:
                st.experimental_set_query_params(sid=st.session_state._sid)
            st.rerun()

        return page

# ── Admin panel ───────────────────────────────────────────────────────────────
def show_admin():
    st.markdown(f"""
    <div class="page-header">
        <h1>🔐 Admin Panel</h1>
        <p>Manage user accounts, access roles, and view system audit logs</p>
    </div>
    """, unsafe_allow_html=True)

    tab_users, tab_audit = st.tabs(["👥 User Management", "📋 System Audit Log"])

    with tab_users:
        users = load_users()
        col_left, col_right = st.columns([1.4, 1])

        with col_left:
            st.markdown('<div class="section-sep"><span>Current Users</span></div>', unsafe_allow_html=True)
            for uname, udata in users.items():
                role_cls  = "admin" if udata["role"] == "admin" else "user"
                initials  = uname[:2].upper()
                is_self   = uname == st.session_state.username
                col1, col2, col3 = st.columns([5, 1.2, 1])
                with col1:
                    st.markdown(f"""
                    <div class="user-row">
                        <div class="u-avatar">{initials}</div>
                        <div style="flex:1">
                            <div class="u-name">{uname} {"(you)" if is_self else ""}</div>
                        </div>
                        <span class="u-role {role_cls}">{udata["role"]}</span>
                    </div>""", unsafe_allow_html=True)
                with col2:
                    if not is_self:
                        if st.button("👤 Login As", key=f"imp_{uname}", use_container_width=True, help=f"Switch to {uname}'s account"):
                            # Save real admin credentials before switching
                            st.session_state.admin_username = st.session_state.username
                            st.session_state.admin_role     = st.session_state.role
                            # Switch to the target user
                            st.session_state.username = uname
                            st.session_state.role     = udata["role"]
                            st.session_state.page     = "report"
                            st.session_state.analysis = None
                            st.rerun()
                    else:
                        st.write("")
                with col3:
                    if not is_self:
                        if st.button("Delete", key=f"del_{uname}", use_container_width=True):
                            del users[uname]; save_users(users); st.rerun()
                    else:
                        st.write("")

        with col_right:
            st.markdown('<div class="section-sep"><span>Add / Update User</span></div>', unsafe_allow_html=True)
            with st.form("add_user"):
                new_user = st.text_input("Username")
                new_pass = st.text_input("Password", type="password")
                new_role = st.selectbox("Role", ["user","admin"])
                if st.form_submit_button("Save User", use_container_width=True, type="primary"):
                    if not new_user or not new_pass:
                        st.error("Username and password are required.")
                    else:
                        users[new_user] = {"hash": hash_password(new_pass), "role": new_role}
                        save_users(users); st.success(f"User '{new_user}' saved."); st.rerun()

    with tab_audit:
        st.markdown('<div class="section-sep"><span>Recent Activities</span></div>', unsafe_allow_html=True)
        log = load_json_file(AUDIT_FILE, [])
        if not log:
            st.info("No audit logs found.")
        else:
            import pandas as pd
            df_audit = pd.DataFrame(log[::-1]) # Reverse for newest first
            st.dataframe(df_audit, use_container_width=True, hide_index=True)
            if st.button("🗑 Clear Audit Log"):
                save_json_file(AUDIT_FILE, [])
                st.rerun()

    st.markdown('<br>', unsafe_allow_html=True)
    if st.button("← Back to Report Generator"):
        st.session_state.page = "report"; st.rerun()

# ── Help page ─────────────────────────────────────────────────────────────────
def show_help():
    st.markdown("""
    <div class="page-header">
        <h1>ℹ️ Help & Documentation</h1>
        <p>Learn how to use the Fee Defaulter Report Generator</p>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("""
    ### Expected Excel Columns

    The tool auto-detects your column names — no renaming required.

    | Field | Accepted column names |
    |---|---|
    | Student | Student, Student Name, Child Name |
    | Parent | Parent Name, Father Name, Guardian Name |
    | Class | Class, Grade, Section |
    | Total Fee | Total Fee, Amount Due, Fee |
    | Paid | Paid, Amount Paid, Received |
    | Pending | Pending, Balance, Outstanding |
    | Phone | Phone, Mobile, Contact |
    | Email | Email, Email Address |
    | Month | Month, Fee Month |
    | Due Date | Due Date, Last Date |

    > If your sheet already has a **Pending / Balance / Outstanding** column, that value is used directly.
    > Otherwise, Pending is calculated as **Total Fee − Paid**.

    ### Tips
    - Upload **multiple files** to combine data across sheets or months.
    - Use **Merge duplicates** to combine the same student's records across files.
    - The **Urdu toggle** generates letters and WhatsApp messages in Urdu.
    - Download the **sample file** to see the expected format before uploading your own.
    """)

# ── Main report page ───────────────────────────────────────────────────────────
def show_report():
    st.markdown("""
    <div class="page-header">
        <h1>📋 Fee Defaulter Report Generator</h1>
        <p>Upload your fee collection workbook to generate defaulter reports and reminder letters</p>
    </div>
    """, unsafe_allow_html=True)

    # ── Settings panel ────────────────────────────────────────────────────────
    with st.expander("⚙️  Report Settings", expanded=True):
        c1, c2 = st.columns(2)
        app_settings = ensure_app_settings()
        school_name    = c1.text_input("School / Institute Name", value=app_settings.get("school_name", "Lahore Grammar School Accounts Office"))
        month_name     = c2.text_input("Fee Month", value=st.session_state.get("source_month", ""), placeholder="e.g. June 2026 — leave blank to auto-detect")
        c3, c4, c5 = st.columns(3)
        tone_opts = ["polite","firm","final"]
        default_tone = app_settings.get("default_tone", "polite")
        dup_opts = ["merge","separate"]
        default_dup = app_settings.get("duplicate_mode", "merge")
        tone           = c3.selectbox("Letter Tone", ["polite","firm","final"],
                                       index=tone_opts.index(default_tone) if default_tone in tone_opts else 0,
                                       format_func={"polite":"🟢  Polite","firm":"🟠  Firm","final":"🔴  Final Warning"}.get)
        duplicate_mode = c4.selectbox("Duplicate Handling", ["merge","separate"],
                                       index=dup_opts.index(default_dup) if default_dup in dup_opts else 0,
                                       format_func={"merge":"Merge rows","separate":"Keep separate"}.get)
        use_urdu       = c5.checkbox("Generate in Urdu 🇵🇰", value=app_settings.get("use_urdu", False))
        currency       = current_currency()

    st.session_state.settings = {
        "school_name": school_name,
        "tone": tone,
        "use_urdu": use_urdu,
        "currency": currency,
    }

    if st.session_state.get("analysis") and st.session_state.get("source_files"):
        last_month = st.session_state.analysis.get("month_override", st.session_state.get("source_month", ""))
        last_dup = st.session_state.analysis.get("duplicate_mode", app_settings.get("duplicate_mode", "merge"))
        if month_name != last_month or duplicate_mode != last_dup:
            with st.spinner("Applying updated report settings to your uploaded records..."):
                try:
                    analysis = analyze_files(st.session_state.source_files, month_name, duplicate_mode)
                    analysis["tone"] = tone
                    analysis["use_urdu"] = use_urdu
                    st.session_state.analysis = analysis
                    st.session_state.source_month = month_name
                except Exception as e:
                    st.error(f"Could not apply updated settings to the current records: {e}")

    # ── Upload section ────────────────────────────────────────────────────────
    st.markdown('<div class="section-sep"><span>Upload Workbook</span></div>', unsafe_allow_html=True)
    uploaded_files = st.file_uploader(
        "Drag and drop your .xlsx / .xlsm files here, or click to browse",
        type=["xlsx","xlsm"], accept_multiple_files=True,
        label_visibility="visible",
    )

    col_gen, col_sample = st.columns([2, 1])
    run_gen    = col_gen.button("🚀  Generate Report", type="primary", use_container_width=True)
    run_sample = col_sample.button("📥  Try Sample File", use_container_width=True)

    # ── Build & analyse ───────────────────────────────────────────────────────
    files_data = []
    if run_sample:
        files_data = [("sample_fee_workbook.xlsx", build_sample_workbook_bytes())]
    elif run_gen:
        if not uploaded_files:
            st.markdown('<div class="banner warn"><span class="banner-icon">⚠️</span> Please upload at least one workbook first, or try the sample file.</div>', unsafe_allow_html=True)
        else:
            for f in uploaded_files:
                files_data.append((f.name, f.read()))

    if files_data:
        with st.spinner("Reading workbook and computing defaulters…"):
            try:
                analysis = analyze_files(files_data, month_name, duplicate_mode)
                analysis["tone"] = tone; analysis["use_urdu"] = use_urdu
                st.session_state.analysis = analysis
                st.session_state.source_files = files_data
                st.session_state.source_month = month_name
                st.session_state.settings = {"school_name": school_name, "tone": tone, "use_urdu": use_urdu, "currency": currency}
            except Exception as e:
                st.error(f"Could not process file(s): {e}")

    analysis = st.session_state.analysis
    if not analysis:
        st.markdown('<div class="banner info"><span class="banner-icon">💡</span> Upload your fee Excel file and click <strong>Generate Report</strong> to get started. You can also try the sample file.</div>', unsafe_allow_html=True)
        return

    settings    = st.session_state.settings
    school_name = settings["school_name"]
    tone        = settings["tone"]
    use_urdu    = settings["use_urdu"]
    currency    = settings.get("currency", current_currency())

    if analysis["warnings"]:
        wcount = len(analysis["warnings"])
        with st.expander(f"⚠️  {wcount} duplicate student{'s' if wcount>1 else ''} detected across uploaded files — click to review", expanded=False):
            st.markdown(
                "<div style='font-size:0.82rem;color:#92400E;line-height:1.8;'>"
                + "<br>".join(f"• {w}" for w in analysis["warnings"])
                + "</div>",
                unsafe_allow_html=True
            )
            st.caption("These students appear in multiple files. If you used 'Merge rows', their fees have been combined into one record.")

    # ── Summary metrics ───────────────────────────────────────────────────────
    st.markdown('<div class="section-sep"><span>Summary & Alerts</span></div>', unsafe_allow_html=True)
    alerts = alert_counts(analysis)
    if any(v > 0 for v in alerts.values()):
        alert_cols = st.columns(len([v for v in alerts.values() if v > 0]))
        curr_col = 0
        if alerts["missing_phone"] > 0:
            alert_cols[curr_col].markdown(f"""<div class="banner warn" style="margin-bottom:15px"><span class="banner-icon">📞</span> <strong>{alerts['missing_phone']}</strong> students missing phone numbers</div>""", unsafe_allow_html=True)
            curr_col += 1
        if alerts["invalid_phone"] > 0:
            alert_cols[curr_col].markdown(f"""<div class="banner warn" style="margin-bottom:15px"><span class="banner-icon">🚫</span> <strong>{alerts['invalid_phone']}</strong> invalid phone formats found</div>""", unsafe_allow_html=True)
            curr_col += 1
        if alerts["duplicates"] > 0:
            alert_cols[curr_col].markdown(f"""<div class="banner info" style="margin-bottom:15px"><span class="banner-icon">👥</span> <strong>{alerts['duplicates']}</strong> duplicate student names identified</div>""", unsafe_allow_html=True)
            curr_col += 1
        if alerts["overdue_30"] > 0:
            alert_cols[curr_col].markdown(f"""<div class="banner warn" style="margin-bottom:15px; background:#FEF2F2; border-color:#FCA5A5; color:#991B1B;"><span class="banner-icon">⏳</span> <strong>{alerts['overdue_30']}</strong> students overdue by 30+ days</div>""", unsafe_allow_html=True)
            curr_col += 1

    t_exp = analysis["total_expected"]
    t_col = analysis["total_collected"]
    t_pen = analysis["total_pending"]
    crate = (t_col / t_exp * 100) if t_exp > 0 else 0
    def_pct = (len(analysis["defaulters"]) / analysis["student_count"] * 100) if analysis["student_count"] else 0

    m1, m2, m3, m4, m5 = st.columns(5)
    cards = [
        (m1, "Total Students",   str(analysis["student_count"]),  "blue",   f"{len(analysis['defaulters'])} defaulters"),
        (m2, "Defaulters",       str(len(analysis["defaulters"])), "red",    f"{def_pct:.0f}% of students"),
        (m3, "Total Expected",   fmt_amount(t_exp, currency),     "indigo",  "Fee target"),
        (m4, "Collected",        fmt_amount(t_col, currency),     "green",  f"{crate:.1f}% collection rate"),
        (m5, "Total Pending",    fmt_amount(t_pen, currency),     "amber",  "Outstanding dues"),
    ]
    for col, label, val, color, sub in cards:
        col.markdown(f"""
        <div class="metric-card {color}">
            <div class="metric-label">{label}</div>
            <div class="metric-value">{val}</div>
            <div class="metric-sub">{sub}</div>
        </div>""", unsafe_allow_html=True)

    st.markdown(f"""
    <div style="margin:8px 0 4px;">
        <div class="progress-label"><span>Collection Progress</span><span>{crate:.1f}%</span></div>
        <div class="progress-wrap"><div class="progress-fill" style="width:{min(crate,100):.1f}%"></div></div>
    </div>""", unsafe_allow_html=True)

    # ── Defaulters table ──────────────────────────────────────────────────────
    st.markdown('<div class="section-sep"><span>Defaulters List</span></div>', unsafe_allow_html=True)
    if not analysis["defaulters"]:
        st.markdown('<div class="banner empty">🎉 No defaulters found — all fees appear to be cleared!</div>', unsafe_allow_html=True)
    else:
        import pandas as pd
        rows = []
        for rec in analysis["defaulters"]:
            msg  = whatsapp_msg_for(rec, tone, school_name, use_urdu, currency)
            link = wa_click_link(rec.get("phone",""), msg)
            rows.append({
                "Student":      rec["student"],
                "Parent":       rec["parent"],
                "Class":        rec["class"],
                "Phone":        rec["phone"],
                "Month":        rec["month"],
                "Total":        rec["total"],
                "Paid":         rec["paid"],
                "Pending":      rec["pending"],
                "Due Date":     rec["due_date"],
                "WhatsApp":     link,
            })
        df = pd.DataFrame(rows)
        st.dataframe(df, use_container_width=True, hide_index=True, column_config={
            "WhatsApp": st.column_config.LinkColumn("WhatsApp", display_text="💬 Open"),
            "Total":   st.column_config.NumberColumn(f"Total ({currency})",   format=f"{currency} %,.0f"),
            "Paid":    st.column_config.NumberColumn(f"Paid ({currency})",    format=f"{currency} %,.0f"),
            "Pending": st.column_config.NumberColumn(f"Pending ({currency})", format=f"{currency} %,.0f"),
        }, height=min(400, 50 + 35 * len(rows)))

    # ── Downloads ─────────────────────────────────────────────────────────────
    st.markdown('<div class="section-sep"><span>Download Reports</span></div>', unsafe_allow_html=True)
    with st.spinner("Preparing files…"):
        excel_bytes = write_report_bytes(analysis, school_name, tone, use_urdu, currency)
        html_bytes  = write_letters_bytes(analysis, school_name, tone, use_urdu, currency)
        wa_bytes    = write_wa_messages_bytes(analysis, school_name, tone, use_urdu, currency)
        csv_bytes   = write_csv_bytes(analysis, school_name, tone, use_urdu, currency)
        class_zip   = class_wise_zip_bytes(analysis, school_name, tone, use_urdu, currency)
        zip_bytes   = build_zip({"fee_defaulter_report.xlsx": excel_bytes,
                                  "reminder_letters.html": html_bytes,
                                  "whatsapp_messages.txt": wa_bytes,
                                  "defaulters.csv": csv_bytes})

    dc1, dc2, dc3, dc4, dc5, dc6 = st.columns(6)
    dc1.download_button("📊 Excel", excel_bytes, "fee_report.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
    dc2.download_button("📄 Letters", html_bytes,  "letters.html",    "text/html",     use_container_width=True)
    dc3.download_button("💬 WhatsApp", wa_bytes,    "messages.txt",    "text/plain",    use_container_width=True)
    dc4.download_button("📋 CSV", csv_bytes,   "defaulters.csv",           "text/csv",      use_container_width=True)
    dc5.download_button("🗂️ Class ZIP", class_zip, "class_wise_reports.zip", "application/zip", use_container_width=True)
    dc6.download_button("📦 All (ZIP)", zip_bytes,   "all_outputs.zip","application/zip",use_container_width=True)

    # ── Letter preview ────────────────────────────────────────────────────────
    if analysis["defaulters"]:
        st.markdown('<div class="section-sep"><span>Letter Preview</span></div>', unsafe_allow_html=True)
        col_sel, col_wa = st.columns([3, 1])
        chosen = col_sel.selectbox("Preview letter for:", [r["student"] for r in analysis["defaulters"]], label_visibility="collapsed")
        rec = next(r for r in analysis["defaulters"] if r["student"] == chosen)
        msg  = whatsapp_msg_for(rec, tone, school_name, use_urdu, currency)
        link = wa_click_link(rec.get("phone",""), msg)
        
        if link:
            if col_wa.button("📱 Prepare WhatsApp", use_container_width=True, type="primary"):
                # Track message history
                history = st.session_state.get("message_history", {})
                key = record_key(rec)
                history.setdefault(key, []).append({
                    "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "tone": tone,
                })
                st.session_state.message_history = history
                save_persistent_state()
                st.session_state.last_wa_link = link
                st.rerun()
            
            if st.session_state.get("last_wa_link") == link:
                st.markdown(f"""
                <div style="margin-top:10px; padding:15px; background:#F0FDF4; border:1px solid #BBF7D0; border-radius:10px; text-align:center;">
                    <p style="color:#166534; font-size:0.85rem; margin-bottom:10px;">✅ Link generated and logged in history!</p>
                    <a href="{link}" target="_blank" style="background:#25D366; color:white; padding:10px 20px; border-radius:8px; text-decoration:none; font-weight:700; display:inline-block;">🚀 OPEN WHATSAPP CHAT</a>
                </div>
                """, unsafe_allow_html=True)

        letter_text = letter_for(rec, tone, school_name, use_urdu, currency)
        st.markdown(f'<div class="letter-card">{letter_text}</div>', unsafe_allow_html=True)

# ── Analytics page ────────────────────────────────────────────────────────────
def show_analytics():
    st.markdown("""
    <div class="page-header">
        <h1>📈 Analytics</h1>
        <p>Visual breakdown of fee collection performance</p>
    </div>
    """, unsafe_allow_html=True)

    analysis = st.session_state.get("analysis")
    if not analysis:
        st.markdown('<div class="banner info"><span class="banner-icon">💡</span> Generate a report first from <strong>Generate Report</strong> to see analytics here.</div>', unsafe_allow_html=True)
        return

    import pandas as pd
    import plotly.graph_objects as go

    defaulters = analysis["defaulters"]
    records    = analysis["records"]
    currency   = current_currency()
    t_exp  = analysis["total_expected"]
    t_col  = analysis["total_collected"]
    t_pen  = analysis["total_pending"]
    crate  = (t_col / t_exp * 100) if t_exp > 0 else 0
    def_pct = (len(defaulters) / analysis["student_count"] * 100) if analysis["student_count"] else 0

    st.markdown('<div class="section-sep"><span>Key Metrics</span></div>', unsafe_allow_html=True)
    m1, m2, m3, m4 = st.columns(4)
    for col, label, val, color, sub in [
        (m1, "Collection Rate",  f"{crate:.1f}%",        "green",  f"{fmt_amount(t_col, currency)} collected"),
        (m2, "Default Rate",     f"{def_pct:.1f}%",       "red",    f"{len(defaulters)} students"),
        (m3, "Avg Pending/Defaulter", fmt_amount(t_pen / len(defaulters), currency) if defaulters else "N/A", "amber", "per defaulter"),
        (m4, "Total Records",    str(analysis["student_count"]), "blue", "students processed"),
    ]:
        col.markdown(f"""
        <div class="metric-card {color}">
            <div class="metric-label">{label}</div>
            <div class="metric-value">{val}</div>
            <div class="metric-sub">{sub}</div>
        </div>""", unsafe_allow_html=True)

    st.markdown('<div class="section-sep"><span>Class-wise Breakdown</span></div>', unsafe_allow_html=True)
    class_stats = {}
    for r in defaulters:
        c = r["class"] or "Uncategorized"
        class_stats[c] = class_stats.get(c, 0) + r["pending"]
    if class_stats:
        pending_col = f"Pending ({currency})"
        money_format = f"{currency} %,.0f"
        df_class = pd.DataFrame({"Class": list(class_stats.keys()), pending_col: list(class_stats.values())}).sort_values(pending_col, ascending=False)
        
        col_c_bar, col_c_pie = st.columns(2)
        
        with col_c_bar:
            fig_class = go.Figure(go.Bar(
                x=df_class["Class"], y=df_class[pending_col],
                marker_color="#3B82F6",
                text=[fmt_amount(v, currency) for v in df_class[pending_col]],
                textposition="outside",
            ))
            fig_class.update_layout(
                margin=dict(t=20, b=20, l=10, r=10), height=320,
                yaxis_title=pending_col, xaxis_title="Class",
                plot_bgcolor="white", paper_bgcolor="white",
                yaxis=dict(gridcolor="#F1F5F9"),
            )
            st.plotly_chart(fig_class, use_container_width=True, config={"displayModeBar": False})

        with col_c_pie:
            fig_class_pie = go.Figure(go.Pie(
                labels=df_class["Class"],
                values=df_class[pending_col],
                hole=0.4,
                textinfo='percent',
            ))
            fig_class_pie.update_layout(
                margin=dict(t=20, b=20, l=10, r=10), height=320,
                showlegend=True,
                legend=dict(orientation="h", yanchor="bottom", y=-0.2, xanchor="center", x=0.5)
            )
            st.plotly_chart(fig_class_pie, use_container_width=True, config={"displayModeBar": False})

        st.dataframe(df_class, use_container_width=True, hide_index=True, column_config={
            pending_col: st.column_config.NumberColumn(pending_col, format=money_format)
        })

    st.markdown('<div class="section-sep"><span>Payment Status Distribution</span></div>', unsafe_allow_html=True)
    paid_count    = len([r for r in records if not r.get("is_defaulter")])
    pending_count = len(defaulters)
    
    col_bar, col_pie = st.columns(2)
    
    with col_bar:
        fig_status = go.Figure(go.Bar(
            x=["Paid", "Pending"],
            y=[paid_count, pending_count],
            marker_color=["#22C55E", "#EF4444"],
            text=[str(paid_count), str(pending_count)],
            textposition="outside",
        ))
        fig_status.update_layout(
            margin=dict(t=20, b=20, l=10, r=10), height=300,
            yaxis_title="Number of Students",
            plot_bgcolor="white", paper_bgcolor="white",
            yaxis=dict(gridcolor="#F1F5F9"),
        )
        st.plotly_chart(fig_status, use_container_width=True, config={"displayModeBar": False})

    with col_pie:
        fig_pie = go.Figure(go.Pie(
            labels=["Paid", "Pending"],
            values=[paid_count, pending_count],
            marker=dict(colors=["#22C55E", "#EF4444"]),
            hole=0.4,
            textinfo='percent+label',
        ))
        fig_pie.update_layout(
            margin=dict(t=20, b=20, l=10, r=10), height=300,
            showlegend=False,
        )
        st.plotly_chart(fig_pie, use_container_width=True, config={"displayModeBar": False})

    st.markdown('<div class="section-sep"><span>Top 10 Highest Pending</span></div>', unsafe_allow_html=True)
    top10 = sorted(defaulters, key=lambda r: r["pending"], reverse=True)[:10]
    if top10:
        pending_col = f"Pending ({currency})"
        money_format = f"{currency} %,.0f"
        df_top = pd.DataFrame([{"Student": r["student"], "Class": r["class"], pending_col: r["pending"]} for r in top10])
        fig_top = go.Figure(go.Bar(
            x=df_top[pending_col], y=df_top["Student"],
            orientation="h",
            marker_color="#8B5CF6",
            text=[fmt_amount(v, currency) for v in df_top[pending_col]],
            textposition="outside",
        ))
        fig_top.update_layout(
            margin=dict(t=20, b=20, l=10, r=10), height=max(260, 36 * len(top10)),
            xaxis_title=pending_col,
            plot_bgcolor="white", paper_bgcolor="white",
            xaxis=dict(gridcolor="#F1F5F9"),
            yaxis=dict(autorange="reversed"),
        )
        st.plotly_chart(fig_top, use_container_width=True, config={"displayModeBar": False})
        st.dataframe(df_top, use_container_width=True, hide_index=True, column_config={
            pending_col: st.column_config.NumberColumn(pending_col, format=money_format)
        })


# ── Student Directory page ─────────────────────────────────────────────────────
def show_student_directory():
    st.markdown("""
    <div class="page-header">
        <h1>👥 Student Directory</h1>
        <p>Browse, filter, and edit student records manually after upload</p>
    </div>
    """, unsafe_allow_html=True)

    analysis = st.session_state.get("analysis")
    if not analysis:
        st.markdown('<div class="banner info"><span class="banner-icon">💡</span> Generate a report first from <strong>Generate Report</strong> to view the directory.</div>', unsafe_allow_html=True)
        return

    import pandas as pd
    records = analysis["records"]
    currency = current_currency()
    followups = st.session_state.get("followups", {})

    st.markdown('<div class="section-sep"><span>Filters</span></div>', unsafe_allow_html=True)
    f1, f2, f3, f4 = st.columns(4)
    search = f1.text_input("🔍 Search", placeholder="Name, parent, or class")
    
    classes = sorted(list(set(r.get("class", "") for r in records if r.get("class"))))
    class_filter = f2.selectbox("Class Filter", ["All"] + classes)
    
    status_filter = f3.selectbox("Payment Status", ["All", "Pending", "Paid"])
    
    overdue_opts = ["All", "1-7 days", "8-15 days", "16-30 days", "30+ days", "No due date", "Not overdue"]
    overdue_filter = f4.selectbox("Overdue Filter", overdue_opts)

    filtered = records
    if search:
        sl = search.lower()
        filtered = [r for r in filtered if sl in r["student"].lower() or sl in r["parent"].lower() or sl in r["class"].lower()]
    if class_filter != "All":
        filtered = [r for r in filtered if r.get("class") == class_filter]
    if status_filter != "All":
        filtered = [r for r in filtered if r.get("status","") == status_filter]
    if overdue_filter != "All":
        filtered = [r for r in filtered if overdue_bucket(r) == overdue_filter]

    st.caption(f"Showing {len(filtered)} of {len(records)} records")

    tab_dir, tab_hist = st.tabs(["📋 Student Records", "📜 Communication History"])

    with tab_dir:
        # ── Table display ──
        rows = []
        for r in filtered:
            key = record_key(r)
            fup = followups.get(key, "Not Contacted")
            
            # Validation highlights
            phone = str(r["phone"]).strip()
            phone_val = phone
            if not phone:
                phone_val = "⚠️ MISSING"
            elif not valid_phone(phone):
                phone_val = f"❌ {phone}"

            rows.append({
                "key":      key,
                "Student":  r["student"],
                "Parent":   r["parent"],
                "Class":    r["class"],
                "Phone":    phone_val,
                "Month":    r["month"],
                "Total":    r["total"],
                "Paid":     r["paid"],
                "Pending":  r["pending"],
                "Status":   r.get("status",""),
                "Overdue":  overdue_bucket(r),
                "Follow-up": fup,
            })
        df = pd.DataFrame(rows)
        
        # We use a custom display logic to allow editing
        if not filtered:
            st.info("No records match your filters.")
        else:
            edited_df = st.data_editor(
                df, 
                use_container_width=True, 
                hide_index=True, 
                column_config={
                    "key": None, # hide key
                    "Total":   st.column_config.NumberColumn(f"Total ({currency})",   format=f"{currency} %,.0f", disabled=True),
                    "Paid":    st.column_config.NumberColumn(f"Paid ({currency})",    format=f"{currency} %,.0f"),
                    "Pending": st.column_config.NumberColumn(f"Pending ({currency})", format=f"{currency} %,.0f"),
                    "Follow-up": st.column_config.SelectboxColumn("Follow-up", options=FOLLOW_UP_STATUSES),
                    "Status": st.column_config.SelectboxColumn("Status", options=["Paid", "Pending"]),
                    "Overdue": st.column_config.TextColumn("Overdue", disabled=True),
                    "Student": st.column_config.TextColumn("Student", disabled=True),
                    "Parent": st.column_config.TextColumn("Parent", disabled=True),
                    "Class": st.column_config.TextColumn("Class", disabled=True),
                    "Phone": st.column_config.TextColumn("Phone", disabled=True, help="Missing or invalid numbers are flagged"),
                    "Month": st.column_config.TextColumn("Month", disabled=True),
                },
                height=min(600, 100 + 35 * len(rows)),
                key="student_editor"
            )

            if st.button("💾 Save Changes to Directory", type="primary"):
                # Sync changes back to st.session_state.analysis["records"] and followups
                changed = False
                for i, row in edited_df.iterrows():
                    orig_key = row["key"]
                    # Find original record in analysis
                    for r in records:
                        if record_key(r) == orig_key:
                            if r["paid"] != row["Paid"] or r["pending"] != row["Pending"] or r["status"] != row["Status"]:
                                r["paid"] = row["Paid"]
                                r["pending"] = row["Pending"]
                                r["status"] = row["Status"]
                                r["is_defaulter"] = (r["pending"] > 0 or r["status"] == "Pending")
                                changed = True
                            
                            current_fup = followups.get(orig_key, "Not Contacted")
                            if current_fup != row["Follow-up"]:
                                followups[orig_key] = row["Follow-up"]
                                changed = True
                            break
                
                if changed:
                    st.session_state.followups = followups
                    recalc_analysis(st.session_state.analysis)
                    save_persistent_state()
                    log_audit("Update directory", f"Modified student records manually")
                    st.success("Changes saved successfully!")
                    st.rerun()

    with tab_hist:
        st.markdown('<div class="section-sep"><span>WhatsApp Message Logs</span></div>', unsafe_allow_html=True)
        history = st.session_state.get("message_history", {})
        if not history:
            st.info("No messages have been sent yet.")
        else:
            hist_rows = []
            for key, logs in history.items():
                # Find student name from key (brute force)
                sname = "Unknown"
                for r in records:
                    if record_key(r) == key:
                        sname = r["student"]
                        break
                for l in logs:
                    hist_rows.append({
                        "Student": sname,
                        "Sent At": l["time"],
                        "Tone": l["tone"].title(),
                    })
            df_hist = pd.DataFrame(hist_rows).sort_values("Sent At", ascending=False)
            st.dataframe(df_hist, use_container_width=True, hide_index=True)
            if st.button("🗑 Clear History"):
                st.session_state.message_history = {}
                save_persistent_state()
                st.rerun()


# ── Message Templates page ─────────────────────────────────────────────────────
def show_message_templates():
    st.markdown("""
    <div class="page-header">
        <h1>✉️ Message Templates</h1>
        <p>Preview and customize WhatsApp and letter message templates</p>
    </div>
    """, unsafe_allow_html=True)

    tmpl = ensure_custom_templates()
    tab1, tab2 = st.tabs(["💬 WhatsApp Messages", "📄 Letter Templates"])

    with tab1:
        st.markdown("**Variables you can use:** `{student}`, `{parent}`, `{amount}`, `{class}`, `{school}`, `{month}`")
        for key, label in [("polite_wa","🟢 Polite"), ("firm_wa","🟠 Firm"), ("final_wa","🔴 Final Warning")]:
            st.markdown(f"**{label}**")
            tmpl[key] = st.text_area(label, value=tmpl[key], key=f"edit_{key}", label_visibility="collapsed", height=80)
        if st.button("💾 Save WhatsApp Templates", type="primary"):
            st.session_state.custom_templates = tmpl
            save_persistent_state()
            log_audit("Update templates", "Modified WhatsApp message templates")
            st.success("WhatsApp templates saved!")
            st.rerun()

    with tab2:
        st.markdown("**Variables you can use:** `{student}`, `{parent}`, `{amount}`, `{class}`, `{school}`, `{month}`")
        for key, label in [("polite_letter","🟢 Polite"), ("firm_letter","🟠 Firm"), ("final_letter","🔴 Final Warning")]:
            st.markdown(f"**{label}**")
            tmpl[key] = st.text_area(label, value=tmpl[key], key=f"edit_{key}", label_visibility="collapsed", height=130)
        if st.button("💾 Save Letter Templates", type="primary"):
            st.session_state.custom_templates = tmpl
            save_persistent_state()
            log_audit("Update templates", "Modified letter templates")
            st.success("Letter templates saved!")

    st.markdown('<div class="section-sep"><span>Live Preview</span></div>', unsafe_allow_html=True)
    preview_vars = _SafeTemplateVars({"student": "Ali Khan", "parent": "Mr. Khan", "amount": fmt_amount(10000), "class": "Grade 7", "school": "Lahore Grammar School", "month": "June 2026", "due_date": "15 June 2026"})
    tone_preview = st.selectbox("Preview tone", ["polite","firm","final"])
    col_wa, col_lt = st.columns(2)
    with col_wa:
        st.markdown("**WhatsApp Preview**")
        wa_text = render_template_text(tmpl[f"{tone_preview}_wa"], preview_vars)
        st.info(wa_text)
    with col_lt:
        st.markdown("**Letter Preview**")
        lt_text = render_template_text(tmpl[f"{tone_preview}_letter"], preview_vars)
        st.text(lt_text)


# ── Fee Calendar page ──────────────────────────────────────────────────────────
def show_fee_calendar():
    st.markdown("""
    <div class="page-header">
        <h1>📅 Fee Calendar</h1>
        <p>Track fee months, due dates, and upcoming collection cycles</p>
    </div>
    """, unsafe_allow_html=True)

    if "fee_calendar" not in st.session_state:
        st.session_state.fee_calendar = []

    col_form, col_list = st.columns([1, 1.5])

    with col_form:
        st.markdown('<div class="section-sep"><span>Add Fee Month</span></div>', unsafe_allow_html=True)
        months = ["January","February","March","April","May","June","July","August","September","October","November","December"]
        years  = [str(y) for y in range(2024, 2028)]
        c1, c2 = st.columns(2)
        sel_month = c1.selectbox("Month", months, index=datetime.now().month - 1)
        sel_year  = c2.selectbox("Year",  years,  index=years.index(str(datetime.now().year)) if str(datetime.now().year) in years else 0)
        due_day   = st.number_input("Due Day of Month", min_value=1, max_value=31, value=15)
        fee_amount = st.number_input(f"Expected Fee Amount ({current_currency()})", min_value=0, value=0, step=500)
        note      = st.text_input("Note (optional)", placeholder="e.g. Annual fee included")
        if st.button("➕ Add to Calendar", type="primary", use_container_width=True):
            entry = {
                "month": f"{sel_month} {sel_year}",
                "due_date": f"{due_day} {sel_month} {sel_year}",
                "amount": fee_amount,
                "note": note,
                "added": datetime.now().strftime("%d %b %Y"),
            }
            # avoid duplicate months
            st.session_state.fee_calendar = [e for e in st.session_state.fee_calendar if e["month"] != entry["month"]]
            st.session_state.fee_calendar.append(entry)
            st.session_state.fee_calendar.sort(key=lambda e: e["month"])
            save_persistent_state()
            log_audit("Update calendar", f"Added/updated month {entry['month']}")
            st.success(f"Added {sel_month} {sel_year}!")
            st.rerun()

    with col_list:
        st.markdown('<div class="section-sep"><span>Calendar Entries</span></div>', unsafe_allow_html=True)
        cal = st.session_state.fee_calendar
        if not cal:
            st.info("No fee months added yet. Add one on the left.")
        else:
            current_month = datetime.now().strftime("%B %Y")
            for i, entry in enumerate(cal):
                is_current = entry["month"] == current_month
                border_color = "#3B82F6" if is_current else "#E2E8F0"
                badge = " 🔵 <em>current</em>" if is_current else ""
                st.markdown(f"""
                <div style="border:1.5px solid {border_color};border-radius:10px;padding:12px 16px;margin-bottom:8px;background:#FAFBFF;">
                    <div style="font-weight:700;color:#1E293B;font-size:0.9rem;">{entry['month']}{badge}</div>
                    <div style="font-size:0.8rem;color:#64748B;margin-top:2px;">📅 Due: {entry['due_date']} &nbsp;|&nbsp; 💰 {fmt_amount(entry['amount']) if entry['amount'] else 'Not set'}</div>
                    {f"<div style='font-size:0.78rem;color:#94A3B8;margin-top:3px;'>📝 {entry['note']}</div>" if entry['note'] else ""}
                </div>""", unsafe_allow_html=True)
                if st.button("🗑 Remove", key=f"rm_cal_{i}"):
                    st.session_state.fee_calendar.pop(i)
                    save_persistent_state()
                    st.rerun()


# ── Export Center page ─────────────────────────────────────────────────────────
def show_export_center():
    st.markdown("""
    <div class="page-header">
        <h1>🖨️ Export Center</h1>
        <p>Download all generated reports and files in one place</p>
    </div>
    """, unsafe_allow_html=True)

    analysis = st.session_state.get("analysis")
    settings = st.session_state.get("settings", {})
    school_name = settings.get("school_name", "Accounts Office")
    tone        = settings.get("tone", "polite")
    use_urdu    = settings.get("use_urdu", False)
    currency    = settings.get("currency", current_currency())

    if not analysis:
        st.markdown('<div class="banner info"><span class="banner-icon">💡</span> Generate a report first from <strong>Generate Report</strong> to enable downloads here.</div>', unsafe_allow_html=True)
        return

    st.markdown('<div class="section-sep"><span>Available Downloads</span></div>', unsafe_allow_html=True)

    with st.spinner("Preparing all files…"):
        excel_bytes = write_report_bytes(analysis, school_name, tone, use_urdu, currency)
        html_bytes  = write_letters_bytes(analysis, school_name, tone, use_urdu, currency)
        wa_bytes    = write_wa_messages_bytes(analysis, school_name, tone, use_urdu, currency)
        csv_bytes   = write_csv_bytes(analysis, school_name, tone, use_urdu, currency)
        class_zip   = class_wise_zip_bytes(analysis, school_name, tone, use_urdu, currency)
        zip_bytes   = build_zip({
            "fee_defaulter_report.xlsx": excel_bytes,
            "reminder_letters.html": html_bytes,
            "whatsapp_messages.txt": wa_bytes,
            "defaulters.csv": csv_bytes,
            "class_wise_reports.zip": class_zip,
        })

    exports = [
        ("📊 Excel Report",       excel_bytes, "fee_defaulter_report.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
         "Full defaulter report with summary dashboard and class-wise breakdown."),
        ("📄 Reminder Letters",   html_bytes,  "reminder_letters.html",    "text/html",
         "Print-ready reminder letters for each defaulter, one per page."),
        ("💬 WhatsApp Messages",  wa_bytes,    "whatsapp_messages.txt",    "text/plain",
         "Pre-written WhatsApp messages with click-to-send links."),
        ("📋 CSV Export",         csv_bytes,   "defaulters.csv",           "text/csv",
         "Raw defaulter data in CSV format for use in other tools."),
        ("🗂️ Class-wise ZIPs",   class_zip,   "class_wise_reports.zip",  "application/zip",
         "Separate Excel and PDF reports for each class in one ZIP."),
        ("📦 All Files (ZIP)",    zip_bytes,   "fee_defaulter_outputs.zip","application/zip",
         "Everything bundled in a single ZIP archive."),
    ]

    for label, data, filename, mime, desc in exports:
        col_info, col_btn = st.columns([4, 1])
        with col_info:
            st.markdown(f"""
            <div style="padding:12px 16px;border:1px solid #E2E8F0;border-radius:10px;background:#FAFBFF;margin-bottom:8px;">
                <div style="font-weight:600;color:#1E293B;font-size:0.9rem;">{label}</div>
                <div style="font-size:0.78rem;color:#64748B;margin-top:2px;">{desc}</div>
                <div style="font-size:0.74rem;color:#94A3B8;margin-top:1px;">📁 {filename}</div>
            </div>""", unsafe_allow_html=True)
        with col_btn:
            st.markdown("<div style='margin-top:4px;'>", unsafe_allow_html=True)
            st.download_button(f"⬇️ Download", data, filename, mime, use_container_width=True, key=f"exp_{filename}")
            st.markdown("</div>", unsafe_allow_html=True)

    st.markdown('<div class="section-sep"><span>Export Summary</span></div>', unsafe_allow_html=True)
    n_def = len(analysis["defaulters"])
    t_pen = analysis["total_pending"]
    st.markdown(f"""
    <div style="background:#EFF6FF;border:1px solid #BFDBFE;border-radius:10px;padding:16px 20px;">
        <div style="font-size:0.875rem;color:#1E40AF;font-weight:600;">Report covers <strong>{n_def} defaulters</strong> with total pending of <strong>{fmt_amount(t_pen, currency)}</strong></div>
        <div style="font-size:0.78rem;color:#3B82F6;margin-top:4px;">School: {school_name} &nbsp;|&nbsp; Tone: {tone.title()} &nbsp;|&nbsp; Language: {"Urdu 🇵🇰" if use_urdu else "English 🇬🇧"}</div>
    </div>""", unsafe_allow_html=True)


# ── Settings page ──────────────────────────────────────────────────────────────
def show_settings_page():
    st.markdown("""
    <div class="page-header">
        <h1>⚙️ Settings</h1>
        <p>Configure default preferences for your school's reports</p>
    </div>
    """, unsafe_allow_html=True)

    previous_cfg = ensure_app_settings().copy()
    cfg = previous_cfg.copy()

    st.markdown('<div class="section-sep"><span>School Information</span></div>', unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    cfg["school_name"] = c1.text_input("School / Institute Name", value=cfg["school_name"])
    cfg["currency"]    = c2.selectbox("Currency Symbol", ["Rs.", "PKR", "$", "£", "€", "AED"], index=["Rs.", "PKR", "$", "£", "€", "AED"].index(cfg.get("currency","Rs.")))

    st.markdown('<div class="section-sep"><span>Report Defaults</span></div>', unsafe_allow_html=True)
    c3, c4, c5 = st.columns(3)
    tone_opts = ["polite","firm","final"]
    cfg["default_tone"]     = c3.selectbox("Default Letter Tone", tone_opts,
                                            index=tone_opts.index(cfg.get("default_tone","polite")),
                                            format_func={"polite":"🟢 Polite","firm":"🟠 Firm","final":"🔴 Final Warning"}.get)
    dup_opts = ["merge","separate"]
    cfg["duplicate_mode"]   = c4.selectbox("Default Duplicate Handling", dup_opts,
                                            index=dup_opts.index(cfg.get("duplicate_mode","merge")),
                                            format_func={"merge":"Merge rows","separate":"Keep separate"}.get)
    cfg["default_due_day"]  = c5.number_input("Default Due Day", min_value=1, max_value=31, value=int(cfg.get("default_due_day", 15)))

    st.markdown('<div class="section-sep"><span>Language</span></div>', unsafe_allow_html=True)
    cfg["use_urdu"] = st.checkbox("Generate letters and messages in Urdu 🇵🇰 by default", value=cfg.get("use_urdu", False))

    st.markdown('<div class="section-sep"><span>School Branding</span></div>', unsafe_allow_html=True)

    if "logo_editor_src" not in st.session_state:
        st.session_state.logo_editor_src = None  # raw uploaded bytes, kept while editing

    logo_file = st.file_uploader("Upload School Logo (PNG/JPG)", type=["png", "jpg", "jpeg"], key="logo_uploader")
    if logo_file is not None:
        new_bytes = logo_file.getvalue()
        if st.session_state.logo_editor_src != new_bytes:
            st.session_state.logo_editor_src = new_bytes

    if st.session_state.logo_editor_src:
        st.caption("Adjust rotation, then drag the square box below to crop your logo.")

        if st.session_state.get("_reset_logo_rotation"):
            st.session_state.logo_rotation = 0
            st.session_state._reset_logo_rotation = False

        c_rotate, c_reset = st.columns([4, 1])
        rotation = c_rotate.slider("Rotate (degrees)", min_value=-180, max_value=180, value=0, step=1, key="logo_rotation")
        with c_reset:
            st.write("")
            if st.button("↺ Reset", use_container_width=True):
                st.session_state._reset_logo_rotation = True
                st.rerun()

        try:
            raw_img = Image.open(io.BytesIO(st.session_state.logo_editor_src)).convert("RGBA")
            rotated_img = raw_img.rotate(-rotation, expand=True, fillcolor=(255, 255, 255, 0))

            c_crop, c_prev = st.columns([2, 1])
            with c_crop:
                cropped_img = st_cropper(
                    rotated_img,
                    realtime_update=True,
                    box_color="#3B82F6",
                    aspect_ratio=(1, 1),
                    return_type="image",
                    key="logo_cropper",
                )
            with c_prev:
                st.markdown("**Live Preview**")
                preview_size = st.slider("Preview size (px)", 60, 300, 150, step=10, key="logo_preview_size")
                st.image(cropped_img, width=preview_size)

            col_save_logo, col_cancel_logo = st.columns(2)
            with col_save_logo:
                if st.button("✅ Save Logo", type="primary", use_container_width=True):
                    final_img = cropped_img.resize((512, 512), Image.LANCZOS).convert("RGBA")
                    buf = io.BytesIO()
                    final_img.save(buf, format="PNG")
                    b64_logo = f"data:image/png;base64,{base64.b64encode(buf.getvalue()).decode()}"
                    cfg["logo_data"] = b64_logo
                    st.session_state.logo_editor_src = None
                    st.success("Logo saved! Click 'Save Settings' below to make it permanent.")
                    st.rerun()
            with col_cancel_logo:
                if st.button("✖️ Cancel", use_container_width=True):
                    st.session_state.logo_editor_src = None
                    st.rerun()
        except Exception as e:
            st.error(f"Could not load this image for editing: {e}")
            st.session_state.logo_editor_src = None

    st.markdown('<div class="section-sep"><span>Current Logo</span></div>', unsafe_allow_html=True)
    if cfg.get("logo_data"):
        st.image(cfg["logo_data"], caption="Logo Preview", width=150)
        if st.button("🗑 Remove Logo"):
            cfg["logo_data"] = ""
            st.session_state.app_settings = cfg
            save_persistent_state()
            log_audit("Update settings", "Removed school logo")
            st.rerun()
    else:
        st.info("No logo uploaded")

    st.markdown('<div class="section-sep"><span>Data Management</span></div>', unsafe_allow_html=True)
    col_bak, col_res = st.columns(2)
    with col_bak:
        st.markdown("**Backup Application Data**")
        st.caption("Export all settings, users, templates, and calendar as one backup file.")
        st.download_button("📥 Download Backup (.json)", build_backup_bytes(), f"fee_portal_backup_{datetime.now().strftime('%Y%m%d')}.json", "application/json", use_container_width=True)
    with col_res:
        st.markdown("**Restore from Backup**")
        st.caption("Upload a previously exported backup file to restore all data.")
        restore_file = st.file_uploader("Upload backup file", type=["json"], label_visibility="collapsed")
        if restore_file:
            try:
                backup_data = json.load(restore_file)
                if st.button("🚀 Confirm Restore", type="primary", use_container_width=True):
                    restore_backup(backup_data)
                    st.success("Application data restored successfully!")
                    st.rerun()
            except Exception as e:
                st.error(f"Invalid backup file: {e}")

    st.markdown('<div class="section-sep"><span>Import Column Mapper</span></div>', unsafe_allow_html=True)
    st.caption("If your Excel headers are unusual, add them here so the tool can recognize them automatically.")
    custom_aliases = cfg.get("custom_aliases", {})
    if not isinstance(custom_aliases, dict): custom_aliases = {}
    
    selected_field = st.selectbox("Select field to add alias for:", list(ALIASES.keys()), format_func=lambda x: x.title())
    new_alias = st.text_input(f"Add new alias for '{selected_field.title()}'", placeholder="e.g. Student Full Name")
    if st.button("➕ Add Alias"):
        if new_alias.strip():
            field_aliases = custom_aliases.get(selected_field, [])
            if new_alias.strip() not in field_aliases:
                field_aliases.append(new_alias.strip())
                custom_aliases[selected_field] = field_aliases
                cfg["custom_aliases"] = custom_aliases
                st.success(f"Added '{new_alias}' as alias for {selected_field}")
                st.rerun()

    if custom_aliases:
        with st.expander("View / Remove Custom Aliases"):
            for field, aliases in custom_aliases.items():
                if aliases:
                    st.write(f"**{field.title()}:** {', '.join(aliases)}")
                    if st.button(f"Clear {field.title()} Aliases", key=f"clear_{field}"):
                        custom_aliases[field] = []
                        cfg["custom_aliases"] = custom_aliases
                        st.rerun()

    st.markdown('<div class="section-sep"><span>Appearance</span></div>', unsafe_allow_html=True)
    st.info("Theme is fixed to the dark navy sidebar with light content area. Contact your admin to customize further.")

    st.markdown("<br>", unsafe_allow_html=True)
    col_save, col_reset = st.columns([1, 1])
    with col_save:
        if st.button("💾 Save Settings", type="primary", use_container_width=True):
            st.session_state.app_settings = cfg
            sync_active_report_settings(cfg)
            if cfg["duplicate_mode"] != previous_cfg["duplicate_mode"] and st.session_state.get("source_files"):
                analysis = analyze_files(st.session_state.source_files, st.session_state.get("source_month", ""), cfg["duplicate_mode"])
                analysis["tone"] = cfg["default_tone"]
                analysis["use_urdu"] = cfg["use_urdu"]
                st.session_state.analysis = analysis
            save_persistent_state()
            log_audit("Update settings", "Modified application preferences")
            st.success("Settings saved successfully!")
    with col_reset:
        if st.button("🔄 Reset to Defaults", use_container_width=True):
            st.session_state.app_settings = DEFAULT_APP_SETTINGS.copy()
            sync_active_report_settings(st.session_state.app_settings)
            if previous_cfg["duplicate_mode"] != "merge" and st.session_state.get("source_files"):
                analysis = analyze_files(st.session_state.source_files, st.session_state.get("source_month", ""), "merge")
                analysis["tone"] = "polite"
                analysis["use_urdu"] = False
                st.session_state.analysis = analysis
            save_persistent_state()
            log_audit("Reset settings", "Restored default application preferences")
            st.success("Settings reset to defaults.")
            st.rerun()

    st.markdown('<div class="section-sep"><span>Account</span></div>', unsafe_allow_html=True)
    st.markdown(f"""
    <div style="background:#F8FAFC;border:1px solid #E2E8F0;border-radius:10px;padding:16px 20px;">
        <div style="font-size:0.875rem;color:#1E293B;"><strong>Logged in as:</strong> {st.session_state.username}</div>
        <div style="font-size:0.8rem;color:#64748B;margin-top:4px;"><strong>Role:</strong> {st.session_state.role.title()}</div>
    </div>""", unsafe_allow_html=True)


# ── Main app ──────────────────────────────────────────────────────────────────
def show_main():
    st.markdown(THEME_CSS, unsafe_allow_html=True)
    page = render_sidebar()

    # ── Impersonation banner ──────────────────────────────────────────────────
    if st.session_state.admin_username:
        col_msg, col_btn = st.columns([5, 1])
        with col_msg:
            st.markdown(f"""
            <div class="impersonation-banner">
                <div>
                    <span class="imp-tag">👁 Viewing As</span>
                    You are viewing the account of <strong>{st.session_state.username}</strong>
                    &nbsp;·&nbsp; Logged in as admin: <strong>{st.session_state.admin_username}</strong>
                </div>
            </div>""", unsafe_allow_html=True)
        with col_btn:
            if st.button("↩ Return to Admin", use_container_width=True, type="primary"):
                # Restore real admin session
                st.session_state.username = st.session_state.admin_username
                st.session_state.role     = st.session_state.admin_role
                st.session_state.admin_username = ""
                st.session_state.admin_role     = ""
                st.session_state.page     = "admin"
                st.session_state.analysis = None
                st.rerun()

    if st.session_state.page == "admin":
        show_admin()
    elif st.session_state.page == "analytics":
        show_analytics()
    elif st.session_state.page == "students":
        show_student_directory()
    elif st.session_state.page == "templates":
        show_message_templates()
    elif st.session_state.page == "calendar":
        show_fee_calendar()
    elif st.session_state.page == "export":
        show_export_center()
    elif st.session_state.page == "settings":
        show_settings_page()
    elif st.session_state.page == "help":
        show_help()
    else:
        st.session_state.page = "report"
        show_report()

# ── Entry point ───────────────────────────────────────────────────────────────
if not st.session_state.logged_in:
    show_login()
else:
    show_main()

_persist_save()